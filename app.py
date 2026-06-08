import os
URL = "https://trwivebwhsvutljhsvll.supabase.co"
KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InRyd2l2ZWJ3aHN2dXRsamhzdmxsIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzM3NDc2MzgsImV4cCI6MjA4OTMyMzYzOH0.9IT5SjOd1mOeDfYWTdnqe7SIIkc49aDVzF9FlJZGknY"
from flask import Flask, render_template, send_from_directory, request, session, flash, redirect, url_for, jsonify, send_file
from supabase import create_client #Client
import bcrypt
import io
import json
import requests
import pandas as pd
from functools import wraps
from datetime import datetime, date, timedelta
from weasyprint import HTML, CSS
from io import BytesIO
from werkzeug.utils import secure_filename
import traceback
import time
import re
from dotenv import load_dotenv
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import pytz  # Ikiwa unahitaji timezone, au tumia datetime.utcnow()
from flask_sitemapper import Sitemapper

def time_ago(dt):
    if isinstance(dt, str):
        # Badilisha 'Z' kuwa '+00:00' na uweke desimali iwe na tarakimu 6
        dt = dt.replace('Z', '+00:00')
        # Ikiwa kuna nukta katika sehemu ya sekunde, rekebisha urefu wa desimali
        if '.' in dt:
            parts = dt.split('.')
            # parts[0] = tarehe na saa (bila desimali)
            # parts[1] = desimali + timezone (mfano '27395+00:00')
            decimal_tz = parts[1]
            if '+' in decimal_tz:
                decimal, tz = decimal_tz.split('+')
                tz = '+' + tz
            else:
                decimal, tz = decimal_tz, ''
            # Hakikisha desimali ina urefu wa tarakimu 6 (ongeza sifuri mwishoni)
            decimal = decimal.ljust(6, '0')[:6]
            dt = f"{parts[0]}.{decimal}{tz}"
        dt = datetime.fromisoformat(dt)
    elif dt.tzinfo is None:
        dt = dt.replace(tzinfo=pytz.UTC)

    now = datetime.now(pytz.UTC)
    diff = now - dt

    if diff.total_seconds() < 10:
        return "now"
    elif diff.total_seconds() < 60:
        s = int(diff.total_seconds())
        return f"{s} second{'s' if s>1 else ''} ago"
    elif diff.total_seconds() < 3600:
        m = int(diff.total_seconds() // 60)
        return f"{m} minute{'s' if m>1 else ''} ago"
    elif diff.total_seconds() < 86400:
        h = int(diff.total_seconds() // 3600)
        return f"{h} hour{'s' if h>1 else ''} ago"
    elif diff.total_seconds() < 2592000:
        d = int(diff.total_seconds() // 86400)
        return f"{d} day{'s' if d>1 else ''} ago"
    else:
        mo = int(diff.total_seconds() // 2592000)
        return f"{mo} month{'s' if mo>1 else ''} ago"
load_dotenv()

def compute_grade(marks):
    """Kokotoa daraja (A, B, C, D, F, x) kutoka alama."""
    if marks is None or marks == '':
        return 'x'
    try:
        marks = float(marks)
    except (ValueError, TypeError):
        return 'x'
    if marks >= 75:
        return 'A'
    elif marks >= 65:
        return 'B'
    elif marks >= 45:
        return 'C'
    elif marks >= 30:
        return 'D'
    else:
        return 'F'



app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'your-secret-key-here')

sitemapper = Sitemapper()

# Hatua ya 3: Unganisha sitemapper na programu yako (init)
sitemapper.init_app(app)

# Supabase configuration

STORAGE_BUCKET = "duty_reports"
supabase = create_client(URL, KEY) if URL and KEY else None

API_KEY = os.environ.get('31b9e69cac7d42388cd07a65fcaae8c4.Dwp57hqv0rzlfd2N')
MODEL = os.environ.get('AI_MODEL', 'glm-4.6v-flash')
API_URL = os.environ.get('AI_API_URL', 'https://api.z.ai/api/paas/v4/chat/completions')

ZAI_API_KEY = os.environ.get('ZAI_API_KEY', '31b9e69cac7d42388cd07a65fcaae8c4.Dwp57hqv0rzlfd2N')
ZAI_API_URL = os.environ.get('ZAI_API_URL', 'https://api.z.ai/api/paas/v4/chat/completions')  # example
ZAI_MODEL = os.environ.get('ZAI_MODEL', 'glm-4.6v-flash')  # adjust to your model

SYSTEM_PROMPT = """You are a translator for school duty teacher reports. Translate the given Kiswahili text into professional English. Follow these rules:
- Translate exactly what is written, do not add or remove information.
- Use past tense and clear, professional English.
- Keep all names, dates, and numbers unchanged.
- Output only the English translation, preserving the original paragraph structure.
- Do not add any extra comments or explanations."""
# ---------- Authentication decorators ----------
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'logged_in' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

# ---------- Routes for pages ----------

@app.route('/favicon.ico')
def favicon():
    return send_from_directory('static', 'img/favicon.ico')

@sitemapper.include(lastmod="2026-06-05")
@app.route('/')
def landing():
    return render_template('landing.html')

@sitemapper.include(lastmod="2026-06-05")
@app.route('/index')
def index():
    return render_template('index.html')  # landing page (si lazima iwe na sidebar)

@app.route('/dashboard')
@login_required
def dashboard():

    result1 = supabase.table('sms_messages')\
        .select('id, sender, time_sent, sms')\
        .order('time_sent', desc=True)\
        .limit(1)\
        .execute()

    if result1.data:
        latest_sms = result1.data[0]
        # Ongeza uga wa relative_time
        latest_sms['relative_time'] = time_ago(latest_sms['time_sent'])

    result = supabase.rpc('get_student_stats_by_form').execute()
    stats_list = result.data

    stats_dict = {}
    all = 0
    active = 0
    truancy = 0
    for row in stats_list:
        class_total = row['active_total'] + row['inactive_total']
        active_total = row['active_total']
        truancy_total = row['inactive_total']
        all += class_total
        active  += active_total
        truancy += truancy_total

        stats_dict[row['class']] = {
            'active_total': row['active_total'],
            'active_boys': row['active_boys'],
            'active_girls': row['active_girls'],
            'inactive_total': row['inactive_total'],
            'inactive_boys': row['inactive_boys'],
            'inactive_girls': row['inactive_girls']
        }

    return render_template('dashboard.html', stats=stats_dict, all=all, sms=latest_sms, active=active, truancy=truancy)


# Students
@app.route('/students-details')
@login_required
def students_details():
    try:
        response = supabase.table('student_list').select('*').execute()
        students = response.data
    except Exception as e:
        students = []
    return render_template('students-details.html', students=students)

@app.route('/new-students')
@login_required
def new_students():
    return render_template('new-students.html')

# Candidates
@app.route('/candidates-registered')
@login_required
def candidates_registered():
    try:
        response = supabase.table('candidates').select('*').execute()
        candidates = response.data
    except Exception:
        candidates = []
    return render_template('candidates-registered.html', candidates=candidates)

@app.route('/candidates-results')
@login_required
def candidates_results():
    return render_template('candidates-results.html')

@app.route('/mbadala')
def mbadala():
    exam_types = supabase.table('exam_types').select('id, exam_name').execute().data
    selected_exam_id = request.args.get('exam_id', type=int, default=exam_types[0]['id'] if exam_types else None)
    if not selected_exam_id:
        return "No exam types found", 404
    result = supabase.rpc('get_student_results_by_exam', {'p_exam_type_id': selected_exam_id}).execute()
    results = result.data
    return render_template('mbadala.html', results=results, exam_types=exam_types, selected_exam_id=selected_exam_id)

# Marking (CA)
@app.route('/ca-scores')
@login_required
def ca_scores():
    return render_template('ca-scores.html')

# (Tayari unayo supabase client)
# supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

@app.route('/marks-entry')
@login_required
def marks_entry():
    # Pata vigezo kutoka query string
    selected_form = request.args.get('form', 'Form1')
    selected_subject = request.args.get('subject', '')
    selected_exam_type_id = request.args.get('exam_type_id', 'all')
    selected_year = request.args.get('year', type=int, default=datetime.now().year)
    selected_term = request.args.get('term', type=int, default=1)

    # Orodha za dropdown (zina 'all')
    forms = ['all', 'Form1', 'Form2', 'Form3', 'Form4', 'Form5', 'Form6']

    subjects_data = supabase.table('subjects').select('subject_name').execute()
    subjects = ['all'] + [s['subject_name'] for s in subjects_data.data]
    if not selected_subject or selected_subject not in subjects:
        selected_subject = subjects[0] if subjects else 'all'

    exam_types_data = supabase.table('exam_types').select('id, exam_name').execute()
    exam_types = [{'id': 'all', 'exam_name': 'All'}] + exam_types_data.data
    if selected_exam_type_id != 'all':
        selected_exam_type_id = int(selected_exam_type_id)

    years = ['all'] + [datetime.now().year - 1, datetime.now().year, datetime.now().year + 1]
    terms = ['all', 1, 2]

    rows = []
    warning = None

    # Kagua kama kuna 'all'
    if selected_form == 'all' or selected_subject == 'all' or selected_exam_type_id == 'all':
        warning = "⚠️ Tafadhali chagua KIDATO, SOMO na AINA YA MTIHANI."
    else:
        try:
            # Itisha SQL function kwa pigo moja
            result = supabase.rpc('get_marks_entry_data', {
                'p_class': selected_form,
                'p_subject_name': selected_subject,
                'p_exam_type_id': selected_exam_type_id,
                'p_year': selected_year,
                'p_term': selected_term
            }).execute()

            # Matokeo yako tayari katika muundo unaofaa
            for item in result.data:
                rows.append({
                    'student_id': item['student_id'],
                    'student_name': item['student_name'],
                    'sex': item['sex'] or '',
                    'marks': item['marks'],
                    'exam_registration_id': item['exam_registration_id']
                })
        except Exception as e:
            warning = f"Hitilafu wakati wa kupata data: {str(e)}"

    return render_template('marks_entry.html',
                           rows=rows,
                           selected_form=selected_form,
                           selected_subject=selected_subject,
                           selected_exam_type_id=selected_exam_type_id,
                           selected_year=selected_year,
                           selected_term=selected_term,
                           forms=forms,
                           subjects=subjects,
                           exam_types=exam_types,
                           years=years,
                           terms=terms,
                           warning=warning)



import traceback
from flask import request, jsonify, session

@app.route('/save-marks', methods=['POST'])
def save_marks():
    # 1. Authentication: Angalia kama mtumiaji ameingia
    if 'user_id' not in session:
        return jsonify({'status': 'error', 'message': 'Hujaoana. Tafadhali ingia kwanza.'}), 401

    user_id = session['user_id']
    user_role = session.get('role')  # 'Academic', 'Teacher', 'Admin', n.k.

    try:
        data = request.json
        form_class = data.get('form')
        subject_name = data.get('subject')
        exam_type_id = data.get('exam_type_id')
        year = data.get('year')
        term = data.get('term')
        marks_list = data.get('marks', [])

        # Hakikisha hakuna 'all'
        if form_class == 'all' or subject_name == 'all' or exam_type_id == 'all':
            return jsonify({'status': 'error', 'message': 'Huwezi kuhifadhi alama ukiwa umechagua "all".'}), 400

        # ------------------------------------------------------------
        # 2. Ukaguzi wa ruhusa (Role-based)
        # ------------------------------------------------------------
        if user_role == 'Academic':
            # Academic ana ruhusa kamili – hakuna haja ya kuangalia teacher_subjects
            pass
        else:
            # Kwa wengine (Teacher, Admin) – angalia kama amepewa somo hili
            teacher_res = supabase.table('teachers_table').select('id').eq('user_id', user_id).execute()
            if not teacher_res.data:
                return jsonify({'status': 'error', 'message': 'Hujasajiliwa kama mwalimu.'}), 403
            teacher_id = teacher_res.data[0]['id']

            teacher_check = supabase.table('teacher_subjects')\
                .select('id')\
                .eq('teacher_id', teacher_id)\
                .eq('subject_name', subject_name)\
                .eq('form', form_class)\
                .execute()
            if not teacher_check.data:
                return jsonify({'status': 'error', 'message': 'Huna ruhusa: wasiliana na mwenye somo.'}), 403

        # ------------------------------------------------------------
        # 3. Pata subject_id kutoka jina la somo
        # ------------------------------------------------------------
        subject_res = supabase.table('subjects').select('id').eq('subject_name', subject_name).execute()
        if not subject_res.data:
            return jsonify({'status': 'error', 'message': 'Somo halipo.'}), 400
        subject_id = subject_res.data[0]['id']

        # ------------------------------------------------------------
        # 4. Badilisha aina za data
        # ------------------------------------------------------------
        try:
            exam_type_id = int(exam_type_id)
            year = int(year)
            term = int(term)
        except (TypeError, ValueError):
            return jsonify({'status': 'error', 'message': 'Exam type, year, na term lazima ziwe namba.'}), 400

        # ------------------------------------------------------------
        # 5. Tengeneza marks_data kwa ajili ya RPC (JSONB)
        # ------------------------------------------------------------
        marks_data = []
        for item in marks_list:
            student_id = item.get('student_id')
            marks_value = item.get('marks')  # inaweza kuwa None (kufuta alama)
            if student_id is not None:
                marks_data.append({
                    'student_id': int(student_id),
                    'marks': marks_value  # None itakubalika kama null
                })

        # ------------------------------------------------------------
        # 6. Piga Supabase RPC save_marks_bulk
        # ------------------------------------------------------------
        result = supabase.rpc('save_marks_bulk', {
            'p_class': form_class,
            'p_subject_id': subject_id,
            'p_exam_type_id': exam_type_id,
            'p_year': year,
            'p_term': term,
            'marks_data': marks_data
        }).execute()

        # result.data ni JSONB yenye 'status' na 'message'
        if result.data and result.data.get('status') == 'success':
            return jsonify({'status': 'success', 'message': result.data.get('message', 'Alama zimehifadhiwa.')})
        else:
            error_msg = result.data.get('message', 'Hitilafu isiyojulikana') if result.data else 'Hitilafu isiyojulikana'
            return jsonify({'status': 'error', 'message': error_msg}), 500

    except Exception as e:
        traceback.print_exc()
        return jsonify({'status': 'error', 'message': str(e)}), 500



@app.route('/create-excel', methods=['POST'])
@login_required
def create_excel():
    try:
        # ============================================
        # DEBUGGING: Print the received JSON
        # ============================================
        print("\n=== RECEIVED REQUEST ===")
        raw_data = request.get_data(as_text=True)
        print(f"Raw data: {raw_data[:500]}")  # first 500 chars

        data = request.get_json()
        print(f"Parsed JSON type: {type(data)}")
        print(f"Keys: {data.keys() if data else 'None'}")

        if data is None:
            return jsonify({'error': 'No JSON data received'}), 400

        # Safely extract marks_data
        marks_data = data.get('data')
        print(f"marks_data type: {type(marks_data)}")
        print(f"marks_data content: {marks_data[:5] if isinstance(marks_data, list) else marks_data}")
        # ============================================

        # Ikiwa marks_data si list, geuza kuwa list tupu
        if marks_data is None or not isinstance(marks_data, list):
            marks_data = []

        if not marks_data:
            return jsonify({'error': 'Hakuna data ya alama'}), 400

        subject_name = data.get('subject', '')
        exam_type = data.get('exam_type', '')
        year = data.get('year', '')
        form = data.get('form', '')

        # ------------------------------
        # Unda Excel kwa kutumia openpyxl (rahisi)
        # ------------------------------
        wb = openpyxl.Workbook()
        ws = wb.active
        sheet_name = form if form else "Alama"
        ws.title = sheet_name

        # Styles
        header_font = Font(name='Arial', size=12, bold=True)
        title_font = Font(name='Arial', size=14, bold=True)
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Row 1
        ws.merge_cells('A1:E1')
        cell = ws['A1']
        cell.value = "KIBINDU SECONDARY SCHOOL"
        cell.font = title_font
        cell.alignment = center_align
        cell.fill = PatternFill(start_color="1E4A3D", end_color="1E4A3D", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True, size=14)

        # Row 2
        ws.merge_cells('A2:E2')
        cell = ws['A2']
        cell.value = f"{subject_name.upper()} - {exam_type.upper()} EXAMINATION {year}"
        cell.font = header_font
        cell.alignment = center_align

        # Row 3 empty
        ws.row_dimensions[3].height = 10

        # Headers row 4
        headers = ['Serial', 'Student Name', 'Sex', 'Marks', 'Signature']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col_idx, value=header)
            cell.font = header_font
            cell.alignment = center_align
            cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
            cell.border = thin_border

        # Data rows
        for i, record in enumerate(marks_data, start=5):
            # Hakikisha record ni dictionary
            if not isinstance(record, dict):
                print(f"WARNING: record {i} is not dict: {record}")
                continue
            ws.cell(row=i, column=1, value=record.get('Serial', ''))
            ws.cell(row=i, column=2, value=record.get('Name', ''))
            ws.cell(row=i, column=3, value=record.get('Sex', ''))
            ws.cell(row=i, column=4, value=record.get('Marks', ''))
            ws.cell(row=i, column=5, value='')

            ws.cell(row=i, column=2).alignment = left_align
            for col in [1,3,4,5]:
                ws.cell(row=i, column=col).alignment = center_align
            for col in range(1,6):
                ws.cell(row=i, column=col).border = thin_border

        # Fixed column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 6
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 15

        # --------------------------------------------------
        # LINDA SHEET (bila password, lakini ruhusu upya wa safu)
        # --------------------------------------------------
        ws.protection.sheet = True
        # Ruhusu kubadilisha upana wa safu (formatColumns = False inamaanisha hairuhusiwi? Hebu fafanua)
        # Kwa kweli: formatColumns = False inaruhusu kubadilisha upana wa safu?
        # Hapana: Kwenye openpyxl, `formatColumns = True` inazuia kubadilisha upana. Tunataka kuwezesha, hivyo weka `formatColumns = False`.
        ws.protection.formatColumns = False   # Ruhusu kubadilisha upana wa safu
        ws.protection.formatRows = False      # Ruhusu kubadilisha urefu wa safu mlalo (hiari)
        ws.protection.insertColumns = False   # Ruhusu kuingiza safu wima? Hii si lazima
        ws.protection.deleteColumns = False   # Ruhusu kufuta safu wima? Hii si lazima

        # Seli zote zimefuliwa (locked) kwa default. Hivyo, funga kwa hakika:
        for row in ws.iter_rows():
            for cell in row:
                cell.protection = openpyxl.styles.Protection(locked=True)

        # Fungua (unlock) safu D (Marks) na E (Signature) kwa safu za data tu
        for row in range(5, 5 + len(marks_data)):
            ws.cell(row=row, column=4).protection = openpyxl.styles.Protection(locked=False)
            ws.cell(row=row, column=5).protection = openpyxl.styles.Protection(locked=False)
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Marks_{form}_{subject_name}_{exam_type}_{year}.xlsx"
        return send_file(output, download_name=filename, as_attachment=True,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/analyse-marks', methods=['POST'])
@login_required
def analyse_marks():
    try:
        data = request.json
        form_class = data.get('form')
        subject_name = data.get('subject')
        exam_type_id = data.get('exam_type_id')
        year = data.get('year')
        term = data.get('term')

        # Thibitisha hakuna 'all'
        if form_class == 'all' or subject_name == 'all' or exam_type_id == 'all':
            return jsonify({'status': 'error', 'message': 'Tafadhali chagua maalum (si all) kwa uchambuzi.'}), 400

        # Pata subject_id
        subject_res = supabase.table('subjects').select('id').eq('subject_name', subject_name).execute()
        if not subject_res.data:
            return jsonify({'status': 'error', 'message': f'Somo "{subject_name}" halipo.'}), 404
        subject_id = subject_res.data[0]['id']

        # Hakikisha exam_type_id ni integer
        exam_type_id = int(exam_type_id)
        year = int(year)
        term = int(term)

        # Pata wanafunzi wa darasa hilo
        students = supabase.table('student_list').select('id').eq('current_class', form_class).execute()
        if not students.data:
            return jsonify({'status': 'error', 'message': f'Hakuna wanafunzi katika {form_class}.'}), 404

        marks_list = []
        for s in students.data:
            # Tafuta exam_registration
            exam_reg = supabase.table('exam_registration')\
                .select('id')\
                .eq('student_id', s['id'])\
                .eq('subject_id', subject_id)\
                .eq('exam_type_id', exam_type_id)\
                .eq('year', year)\
                .eq('term', term)\
                .eq('class', form_class)\
                .execute()
            if exam_reg.data:
                # Tafuta marks
                marks_res = supabase.table('marks_table')\
                    .select('marks')\
                    .eq('exam_registration_id', exam_reg.data[0]['id'])\
                    .execute()
                if marks_res.data and marks_res.data[0]['marks'] is not None:
                    marks_list.append(marks_res.data[0]['marks'])

        if not marks_list:
            return jsonify({'status': 'error', 'message': 'Hakuna marks, save marks kwanza ndo ufanye analysis.'}), 404

        # Count grades
        grade_counts = {'A': 0, 'B': 0, 'C': 0, 'D': 0, 'F': 0}
        for m in marks_list:
            if m >= 75:
                grade_counts['A'] += 1
            elif m >= 65:
                grade_counts['B'] += 1
            elif m >= 45:
                grade_counts['C'] += 1
            elif m >= 30:
                grade_counts['D'] += 1
            else:  # m >= 0 (since marks are non-negative)
                grade_counts['F'] += 1
        analysis = {
            'total_students': len(marks_list),
            'average': round(sum(marks_list) / len(marks_list), 2),
            'max': max(marks_list),
            'min': min(marks_list),
            'A': grade_counts['A'],
            'B': grade_counts['B'],
            'C': grade_counts['C'],
            'D': grade_counts['D'],
            'F': grade_counts['F']
        }
        return jsonify({'status': 'success', 'analysis': analysis})

    except Exception as e:
        return jsonify({'status': 'error', 'message': f'Internal server error: {str(e)}'}), 500
# Academic
@app.route('/all-teachers')
@login_required
def all_teachers():
    try:
        response = supabase.table('teachers').select('*').execute()
        teachers = response.data
    except Exception:
        teachers = []
    return render_template('all-teachers.html', teachers=teachers)

@app.route('/subject-teachers')
@login_required
def subject_teachers():
    return render_template('subject-teachers.html')

@app.route('/teachers-activity')
@login_required
def teachers_activity():
    return render_template('teachers-activity.html')

@app.route('/class-teachers')
@login_required
def class_teachers():
    return render_template('class-teachers.html')



@app.route('/timetables')
@login_required
def timetables():
    return render_template('timetables.html')

@app.route('/teaching-progress')
@login_required
def teaching_progress():
    return render_template('teaching-progress.html')

# Teacher on duty
@app.route('/week-details')
@login_required
def week_details():
    return render_template('week-details.html')

mbadala
# Centre
@app.route('/school-info')
@login_required
def school_info():
    return render_template('school-info.html')

@app.route('/school-almanac')
@login_required
def school_almanac():
    return render_template('school-almanac.html')

# Mhasibu
@app.route('/food')
@login_required
def food():
    return render_template('food.html')

@app.route('/contributions')
@login_required
def contributions():
    return render_template('contributions.html')

@app.route('/prepo')
def prepo():
    return render_template('prepo.html')
@app.route('/mkuu')
@login_required
def mkuu():
    return render_template('mkuu.html')

@app.route('/academic_teachers', methods=['GET', 'POST'])
@login_required
def academic_teachers():
    # Hakikisha ni Mtaaluma pekee
    if session.get('role') != 'Academic':
        flash("Huna ruhusa ya kufikia ukurasa huu.", "error")
        return redirect(url_for('mtaaluma'))

    # ---------- Fafanua variables kabla ya matumizi ----------
    teachers = []
    available_users = []

    if request.method == 'POST':
        # Ikiwa kuna teacher_id, ni kuhariri; vinginevyo ni kuongeza mwalimu mpya
        teacher_id = request.form.get('teacher_id')

        if teacher_id:
            # ---------- EDIT: Sasisha class_teacher na masomo ----------
            class_teacher = request.form.get('class_teacher', '').strip()
            subject_names = request.form.getlist('subject_name[]')
            subject_forms = request.form.getlist('subject_form[]')

            try:
                # Sasisha class_teacher
                supabase.table('teachers_table').update({
                    'class_teacher': class_teacher
                }).eq('id', teacher_id).execute()

                # Sasisha masomo: futa yote na uingize mapya
                supabase.table('teacher_subjects').delete().eq('teacher_id', teacher_id).execute()
                for subj, form in zip(subject_names, subject_forms):
                    if subj and form:
                        supabase.table('teacher_subjects').insert({
                            'teacher_id': teacher_id,
                            'subject_name': subj.strip(),
                            'form': form.strip()
                        }).execute()

                flash("Taarifa zimehifadhiwa kikamilifu.", "success")
            except Exception as e:
                flash(f"Hitilafu wakati wa kuhariri: {str(e)}", "error")
        else:
            # ---------- ADD NEW TEACHER ----------
            username = request.form.get('username', '').strip()
            full_name = request.form.get('full_name', '').strip()

            if not username or not full_name:
                flash("Tafadhali jaza jina kamili na uchague username.", "error")
                return redirect(url_for('academic_teachers'))

            # Tafuta user_id kutoka username
            user_res = supabase.table('users_table').select('id, role').eq('username', username).execute()
            if not user_res.data:
                flash(f"Username '{username}' haipatikani.", "error")
                return redirect(url_for('academic_teachers'))

            user_id = user_res.data[0]['id']
            user_role = user_res.data[0]['role']

            # Hakikisha mtumiaji huyu bado hajawekwa kwenye teachers_table
            existing_teacher = supabase.table('teachers_table').select('id').eq('user_id', user_id).execute()
            if existing_teacher.data:
                flash(f"Mtumiaji '{username}' tayari ana wasifu wa mwalimu.", "error")
                return redirect(url_for('academic_teachers'))

            # Hakikisha mtumiaji ana role inayofaa
            if user_role not in ['Teacher', 'Academic']:
                flash(f"Mtumiaji '{username}' hana role inayofaa (Teacher/Academic).", "error")
                return redirect(url_for('academic_teachers'))

            # Ingiza rekodi mpya kwenye teachers_table
            try:
                supabase.table('teachers_table').insert({
                    'user_id': user_id,
                    'teacher_name': full_name,
                    'class_teacher': '',
                    'role': user_role,
                    'week_status': 'off_duty'
                }).execute()
                flash(f"Mwalimu '{full_name}' ameongezwa kikamilifu.", "success")
            except Exception as e:
                flash(f"Hitilafu wakati wa kuongeza: {str(e)}", "error")

        return redirect(url_for('academic_teachers'))

    # ---------- GET: Onyesha walimu na watumiaji wanaoweza kuongezwa ----------
    try:
        # Chukua walimu wote
        teachers_res = supabase.table('teachers_table').select('*').execute()
        for t in teachers_res.data:
            # Chukua username kutoka users_table
            user_res = supabase.table('users_table').select('username').eq('id', t['user_id']).execute()
            t['username'] = user_res.data[0]['username'] if user_res.data else 'Unknown'

            # Chukua masomo ya mwalimu kutoka teacher_subjects
            subjects_res = supabase.table('teacher_subjects').select('subject_name, form').eq('teacher_id', t['id']).execute()
            t['subjects_list'] = subjects_res.data if subjects_res.data else []   # daima ni list

            teachers.append(t)

        # Chukua watumiaji wote wenye role Teacher/Academic ambao hawajapewa wasifu bado
        users_res = supabase.table('users_table').select('id, username, role').in_('role', ['Teacher', 'Academic']).execute()
        teacher_user_ids = [t['user_id'] for t in teachers]
        available_users = [u for u in users_res.data if u['id'] not in teacher_user_ids]

    except Exception as e:
        flash(f"Hitilafu ya kupata data: {e}", "error")
        # teachers na available_users tayari zimefafanuliwa kama orodha tupu
    return render_template('academic/teachers2.html', teachers=teachers, available_users=available_users)


@app.route('/teachers', methods=['GET', 'POST'])
@login_required
def teachers():
    # Hakikisha ni Head (superadmin) pekee
    if not session.get('logged_in') or session.get('role') != 'Head':
        flash("Huna ruhusa ya kufikia ukurasa huu.", "error")
        return redirect(url_for('mkuu'))

    if request.method == 'POST':
        # ---------- ADD NEW USER ----------
        username = request.form.get('username', '').strip()
        role = request.form.get('role')   # 'Teacher' au 'Academic'

        # Validation
        if not username:
            flash("Tafadhali ingiza jina la mtumiaji.", "error")
            return redirect(url_for('teachers'))

        if role not in ['Teacher', 'Academic']:
            flash("Role siyo sahihi.", "error")
            return redirect(url_for('teachers'))

        # Check if username already exists
        existing = supabase.table('users_table').select('username').eq('username', username).execute()
        if existing.data:
            flash(f"Jina la mtumiaji '{username}' tayari lipo.", "error")
            return redirect(url_for('teachers'))

        # Default password based on role
        raw_password = "secret1234" if role == 'Teacher' else "secret234"

        # Hash password
        hashed = bcrypt.hashpw(raw_password.encode('utf-8'), bcrypt.gensalt()).decode()

        # Insert into Supabase
        try:
            supabase.table('users_table').insert({
                "username": username,
                "password": hashed,
                "role": role,
                "created_at": datetime.utcnow().isoformat()
            }).execute()
            flash(f"{role} '{username}' ameongezwa. Password yake ya awali: {raw_password}", "success")
        except Exception as e:
            flash(f"Hitilafu wakati wa kuingiza: {str(e)}", "error")

        return redirect(url_for('teachers'))

    # ---------- GET: Show list of users ----------
    response = supabase.table('users_table').select('username, role').in_('role', ['Teacher', 'Academic']).execute()
    users = response.data

    # Debug: angalia kama kuna username None (ili kuzuia BuildError)
    for u in users:
        if u.get('username') is None:
            print(f"Warning: User with role {u.get('role')} has null username!")

    return render_template('teachers.html', users=users)

@app.route('/mtaaluma')
@login_required
def mtaaluma():
    return render_template('mtaaluma.html')
@app.route('/schemes')
def schemes():
    return render_template('schemes.html')
@app.route('/lesson_plan')
def accountant_reports():
    return render_template('lesson_notes.html')
@app.route('/lesson_notes')
def lesson_notes():
    return render_template('lesson_notes.html')
@app.route('/questions')
def questions():
    return render_template('questions.html')
# Documents
@app.route('/sms')
@login_required
def sms():
    return render_template('lesson_notes.html')
@app.route('/profile')
@login_required
def profile():
    return render_template('profile.html')

#from werkzeug.security import generate_password_hash, check_password_hash

@app.route('/change-password', methods=['GET', 'POST'])
def change_password():
    if not session.get('logged_in'):
        return redirect(url_for('login'))

    if request.method == 'POST':
        old = request.form.get('old_password')
        new = request.form.get('new_password')
        confirm = request.form.get('confirm_password')

        if new != confirm:
            flash("Nywila mpya na uthibitisho hazilingani", "error")
            return redirect(url_for('change_password'))

        username = session['username']
        # Tafuta mtumiaji
        res = supabase.table('users_table').select('password').eq('username', username).execute()
        if not res.data:
            flash("Hitilafu, jaribu tena", "error")
            return redirect(url_for('change_password'))

        stored = res.data[0]['password']
        if not bcrypt.checkpw(old.encode('utf-8'), stored.encode('utf-8')):
            flash("Password ya zamani si sahihi", "error")
            return redirect(url_for('change_password'))

        # Sasisha password mpya
        new_hashed = bcrypt.hashpw(new.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        supabase.table('users_table').update({'password': new_hashed}).eq('username', username).execute()
        flash("Password imebadilishwa kikamilifu", "success")
        return redirect(url_for('dashboard'))

    return render_template('change_password.html')
    # GET request – onyesha fomu
    return render_template('change_password.html')


# ---------- Login / Logout ----------
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')


        # Tafuta mtumiaji kwenye Supabase
        response = supabase.table('users_table').select('*').eq('username', username).execute()
        users = response.data

        if users:
            user = users[0]
            stored_hash = user['password']
            if bcrypt.checkpw(password.encode('utf-8'), stored_hash.encode('utf-8')):
                session['logged_in'] = True
                session['username'] = username
                session['role'] = user['role']
                session['user_id'] = user['id']   # supabase inarudisha 'id' kwa jedwali la users_table

                # Elekeza kulingana na role
                if user['role'] == 'Teacher':
                    flash(f"Karibu mwalimu {username}.", "success")
                    return redirect(url_for('dashboard'))
                elif user['role'] == 'Academic':
                    flash(f"Karibu mtaaluma: {username}.", "success")
                    return redirect(url_for('mtaaluma'))
                elif user['role'] == 'Head':
                    flash(f"Karibu Mkuu wa shule Ndg: {username}.", "success")
                    return redirect(url_for('mkuu'))
                else:
                    # role isiyojulikana
                    flash("Role yako haijatambuliwa.", "error")
                    return redirect(url_for('login'))
            else:
                flash("Nywila si sahihi", "error")
        else:
            flash("Jina la mtumiaji halipo", "error")
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('index'))

# ---------- Error handlers ----------
@app.errorhandler(404)
def page_not_found(e):
    return render_template('404.html'), 404
'''
@app.errorhandler(500)
def internal_error(e):
    return render_template('500.html'), 500
'''
@app.errorhandler(500)
def internal_error(error):
    return render_template('500.html'), 500

@app.errorhandler(Exception)
def handle_exception(e):
    # Andika log kwenye faili (hiari)
    app.logger.error(f"Unhandled Exception: {e}")
    return render_template('500.html'), 500
def get_user_balance():
    # Simulate with session or a simple file per user
    # For demo, use session['balance'] or default 200
    return session.get('balance', 200)

def deduct_balance(amount=100):
    current = session.get('balance', 200)
    session['balance'] = current - amount
    return session['balance']



@app.route('/lesson-plan', methods=['GET', 'POST'])
def lesson_plan():
    if request.method == 'POST':
        # Handle form submission if needed (for PDF generation etc.)
        # For now, we only use AJAX
        pass
    # For GET, render the form
    # Get initial data (topics, etc.) from static data
    return render_template('lesson_plan.html')

@app.route('/api/topics', methods=['POST'])
def get_topics():
    """Return topics based on subject (mocked)"""
    data = request.get_json()
    subject = data.get('subject')
    # Mock data – in real app, fetch from database or static mapping
    topics = {
        'Biology': ['Cell Structure', 'Photosynthesis', 'Human Digestive System'],
        'Physics': ['Newton\'s Laws', 'Electricity', 'Optics'],
        'Chemistry': ['Periodic Table', 'Chemical Bonding', 'Acids and Bases'],
        'Mathematics': ['Algebra', 'Geometry', 'Calculus'],
        'English': ['Narrative Writing', 'Poetry Analysis', 'Grammar'],
        'Kiswahili': ['Ushairi', 'Fasihi Simulizi', 'Sarufi']
    }
    return jsonify(topics.get(subject, []))

@app.route('/api/specific-activities', methods=['POST'])
def get_specific_activities():
    """Return specific activities based on topic and subject (mocked)"""
    data = request.get_json()
    subject = data.get('subject')
    topic = data.get('topic')
    # Mock data – could be dynamic
    activities = [
        "Define the concept and identify key components.",
        "Explain the process with examples.",
        "Apply the concept to real-life situations.",
        "Analyze case studies related to the topic."
    ]
    return jsonify(activities)

@app.route('/api/generate-lesson', methods=['POST'])
def generate_lesson():
    """Call AI API to generate lesson plan structure"""
    data = request.get_json()
    class_level = data.get('class_level')
    subject = data.get('subject')
    topic = data.get('topic')
    duration = data.get('duration')
    methods = data.get('methods')
    specific_activity = data.get('specific_activity')
    # Additional fields can be included
    # Build prompt similar to original
    prompt = f"""
    You are an experienced professional teacher and curriculum expert of Tanzania.

    Generate a guided part of lesson plan for a {class_level} class
    using the specific objective below on the topic '{topic}'
    under the subject '{subject}', lasting {duration} minutes.

    Specific objective: {specific_activity}
    Suggested methods for new knowledge teaching activity: {methods}

    IMPORTANT RULES (Follow strictly)!!!
    1. Return ONLY valid JSON.
    2. Do NOT add explanations; summarize each teaching and learning activity and assessment in <=35 words.
    3. Follow the exact JSON TEMPLATE below.
    4. Teaching activity should start with words like 'Leading student to', 'Guiding student to'.
    5. Assessment should describe what the student is able to do at each stage starting with: Student able to ......
    6. Specific objective should follow SMART (Specific, Measurable, Attainable, Realistic, Time-bound).
    7. Introduction stage takes ≈ 13% of {duration} minutes, New knowledge ≈ 38% of {duration} minutes, Design takes ≈ 25% of {duration} minutes, Realization takes ≈ 25% of {duration} minutes

    JSON TEMPLATE:
    {{
      "specific_activity": "",
      "lesson_development": [
        {{"stage": "Introduction","time_minutes": 0,"teaching_activity": "","learning_activity": "","assessment": ""}},
        {{"stage": "New Knowledge","time_minutes": 0,"teaching_activity": "","learning_activity": "","assessment": ""}},
        {{"stage": "Design","time_minutes": 0,"teaching_activity": "","learning_activity": "","assessment": ""}},
        {{"stage": "Realisation","time_minutes": 0,"teaching_activity": "","learning_activity": "","assessment": ""}}
      ]
    }}
    """
    headers = {
        "Authorization": f"Bearer {API_KEY}",
        "Content-Type": "application/json"
    }
    payload = {
        "model": MODEL,
        "messages": [
            {"role": "system", "content": "You are a professional teacher assistant."},
            {"role": "user", "content": prompt}
        ],
        "temperature": 0.3
    }
    try:
        response = requests.post(API_URL, json=payload, headers=headers, timeout=60)
        response.raise_for_status()
        raw = response.json()["choices"][0]["message"]["content"]
        # Attempt to parse JSON from raw string
        # Sometimes the AI may include extra text; we extract the JSON part
        import re
        json_match = re.search(r'\{.*\}', raw, re.DOTALL)
        if json_match:
            lesson_data = json.loads(json_match.group())
        else:
            lesson_data = json.loads(raw)
        return jsonify(lesson_data)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/save-pdf', methods=['POST'])
def save_pdf():
    """Generate PDF from lesson plan data (client will request this after confirmation)"""
    # We'll use weasyprint to generate PDF server-side, or use a client-side library
    # For simplicity, we'll return a link to a generated PDF file or just a placeholder
    data = request.get_json()
    # For now, just return success
    # In real app, generate PDF and send as file
    return jsonify({"status": "success", "message": "PDF generated (simulated)"})

@app.route('/api/check-balance', methods=['GET'])
def check_balance():
    balance = get_user_balance()
    return jsonify({"balance": balance})

@app.route('/api/deduct-balance', methods=['POST'])
def deduct():
    amount = request.json.get('amount', 100)
    balance = get_user_balance()
    if balance >= amount:
        new_balance = deduct_balance(amount)
        return jsonify({"success": True, "balance": new_balance})
    else:
        return jsonify({"success": False, "message": "Insufficient balance"}), 402

@app.route('/delete-user/<username>')
def delete_user(username):
    if not session.get('logged_in') or session.get('role') != 'Head':
        flash("Huna ruhusa ya kufuta watumiaji.", "error")
        return redirect(url_for('login'))

    # Usiruhusu kujifuta mwenyewe
    if username == session['username']:
        flash("Huwezi kujifuta mwenyewe.", "error")
        return redirect(url_for('teachers'))

    # Angalia kama mtumiaji yupo na ana role sahihi
    existing = supabase.table('users_table').select('role').eq('username', username).execute()
    if not existing.data:
        flash("Mtumiaji hapatikani.", "error")
        return redirect(url_for('teachers'))

    role = existing.data[0]['role']
    if role not in ['Teacher', 'Academic']:
        flash("Unaweza kufuta tu walimu na wataaluma.", "error")
        return redirect(url_for('teachers'))

    # Futa mtumiaji
    try:
        supabase.table('users_table').delete().eq('username', username).execute()
        flash(f"Mtumiaji '{username}' amefutwa kikamilifu.", "success")
    except Exception as e:
        flash(f"Hitilafu wakati wa kufuta: {str(e)}", "error")
    return redirect(url_for('teachers'))

@app.route('/edit-user', methods=['POST'])
def edit_user():
    # Hakikisha ni Head (superadmin)
    if not session.get('logged_in') or session.get('role') != 'Head':
        flash("Huna ruhusa ya kuhariri watumiaji.", "error")
        return redirect(url_for('login'))

    original_username = request.form.get('original_username')
    new_username = request.form.get('username', '').strip()
    new_role = request.form.get('role')

    # Validation
    if not new_username:
        flash("Jina la mtumiaji haliwezi kuwa tupu.", "error")
        return redirect(url_for('teachers'))

    if new_role not in ['Teacher', 'Academic']:
        flash("Role siyo sahihi.", "error")
        return redirect(url_for('teachers'))

    # Angalia kama mtumiaji yupo
    existing = supabase.table('users_table').select('*').eq('username', original_username).execute()
    if not existing.data:
        flash("Mtumiaji hapatikani.", "error")
        return redirect(url_for('teachers'))

    # Ikiwa jina limebadilishwa, hakikisha halijachukuliwa na mwingine
    if new_username != original_username:
        check = supabase.table('users_table').select('username').eq('username', new_username).execute()
        if check.data:
            flash(f"Jina la mtumiaji '{new_username}' tayari lipo.", "error")
            return redirect(url_for('teachers'))

    # Sasisha data
    try:
        supabase.table('users_table').update({
            'username': new_username,
            'role': new_role
        }).eq('username', original_username).execute()
        flash(f"Taarifa za '{new_username}' zimehaririwa kikamilifu.", "success")
    except Exception as e:
        flash(f"Hitilafu wakati wa kuhariri: {str(e)}", "error")
    return redirect(url_for('teachers'))

@app.route('/teachers-work')
@login_required
def teachers_work():
    # Ruhusa: Unaweza kuweka kwa role yoyote unayotaka (k.m. Head, Academic, Teacher)
    # Kwa mfano, wote walioingia wanaweza kuona
    teachers = []
    try:
        # Chukua walimu wote kutoka teachers_table
        teachers_res = supabase.table('teachers_table').select('id, teacher_name, other_role, week_status').execute()
        for t in teachers_res.data:
            # Chukua masomo ya mwalimu kutoka teacher_subjects
            subjects_res = supabase.table('teacher_subjects').select('subject_name, form').eq('teacher_id', t['id']).execute()
            subjects_list = subjects_res.data if subjects_res.data else []
            # Unganisha masomo na vidato kwa muundo "Somo (Kidato)"
            subjects_str = ', '.join([f"{subj['subject_name']} ({subj['form']})" for subj in subjects_list])
            t['subjects'] = subjects_str if subjects_str else '-'
            teachers.append(t)
    except Exception as e:
        flash(f"Hitilafu: {e}", "error")
    return render_template('teachers_work.html', teachers=teachers)

@app.route('/api/sms-readers/<int:sms_id>')
@login_required
def sms_readers(sms_id):
    user_id = session.get('user_id')
    if not user_id:
        return jsonify({'error': 'Unauthorized'}), 401
    try:
        res = supabase.table('sms_read_receipts').select('user_id, read_at').eq('sms_id', sms_id).execute()
        readers = []
        for r in (res.data or []):
            # Tafuta username kutoka users_table (badilisha ikiwa unatumia 'username' au 'name')
            user_res = supabase.table('users_table').select('username').eq('id', r['user_id']).execute()
            username = user_res.data[0]['username'] if user_res.data else str(r['user_id'])
            readers.append({'username': username, 'read_at': r['read_at']})
        return jsonify(readers)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/mark-sms-read', methods=['POST'])
@login_required
def mark_sms_read():
    user_id = session.get('user_id')
    if not user_id:
        return jsonify({'error': 'Unauthorized'}), 401
    data = request.get_json()
    sms_id = data.get('sms_id')
    if not sms_id:
        return jsonify({'error': 'sms_id required'}), 400
    try:
        existing = supabase.table('sms_read_receipts').select('id').eq('sms_id', sms_id).eq('user_id', user_id).execute()
        if not existing.data:
            supabase.table('sms_read_receipts').insert({'sms_id': sms_id, 'user_id': user_id}).execute()
        return jsonify({'status': 'ok'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/get-sms')
@login_required
def get_sms():
    user_id = session.get('user_id')
    if not user_id:
        return jsonify({'error': 'User not identified'}), 401
    try:
        # Chukua sms zote, pamoja na idadi ya wasomaji
        sms_res = supabase.table('sms_messages').select('*').order('time_sent', desc=True).execute()
        sms_list = sms_res.data or []

        # Chukua read receipts za user huyu
        read_res = supabase.table('sms_read_receipts').select('sms_id').eq('user_id', user_id).execute()
        read_ids = {r['sms_id'] for r in (read_res.data or [])}

        for sms in sms_list:
            sms['is_read'] = sms['id'] in read_ids
            # Idadi ya wasomaji kwa sms hii
            count_res = supabase.table('sms_read_receipts').select('user_id', count='exact').eq('sms_id', sms['id']).execute()
            sms['read_count'] = count_res.count if hasattr(count_res, 'count') else 0

        return jsonify(sms_list)
    except Exception as e:
        print("Error in /api/get-sms:", e)
        return jsonify({'error': str(e)}), 500

@app.route('/api/send-sms', methods=['POST'])
@login_required
def send_sms():
    # Ruhusu watumiaji wote walioingia (au unaweza kuweka list ya roles)
    if not session.get('logged_in'):
        return jsonify({'error': 'Unauthorized'}), 401
    # (hiari) if session.get('role') not in ['Head', 'Academic', 'Teacher']: return jsonify(...),403
    data = request.get_json()
    sender = data.get('sender')
    sms_text = data.get('sms')
    if not sender or not sms_text:
        return jsonify({'error': 'Missing fields'}), 400
    supabase.table('sms_messages').insert({
        'sender': sender,
        'sms': sms_text,
        'time_sent': datetime.utcnow().isoformat()
    }).execute()
    return jsonify({'status': 'ok'})

@app.route('/subjects', methods=['GET'])
def get_subjects():
    try:
        result = supabase.table('subjects').select('*').execute()
        return jsonify(result.data), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/sync-subjects')
def sync_subjects_page():
    return render_template('sync_subjects.html')

@app.route('/sync-teachers', methods=['POST'])
def sync_teachers():
    try:
        # Itisha Supabase function
        response = supabase.rpc('sync_teacher_names_to_subjects').execute()
        # Angalia kama kuna error kutoka Supabase
        if hasattr(response, 'error') and response.error:
            return jsonify({"error": str(response.error)}), 500
        # Rudisha ujumbe wa mafanikio
        return jsonify({"message": "Success", "data": response.data}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/edit-other-roles', methods=['GET', 'POST'])
@login_required
def edit_other_roles():
    # Ruhusa: Head au Academic pekee (rekebisha kulingana na mfumo wako)
    if session.get('role') not in ['Head', 'Academic']:
        flash("Huna ruhusa ya kuhariri majukumu ya walimu.", "error")
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        teacher_id = request.form.get('teacher_id')
        other_role = request.form.get('other_role', '').strip()
        if teacher_id:
            supabase.table('teachers_table').update({'other_role': other_role}).eq('id', teacher_id).execute()
            flash(f"Other role imehifadhiwa kwa mwalimu.", "success")
        return redirect(url_for('edit_other_roles'))

    # GET: onyesha orodha ya walimu
    teachers = supabase.table('teachers_table').select('id, teacher_name, other_role').execute()
    teachers_data = teachers.data if teachers.data else []

    # Orodha ya majukumu yanayopatikana (kwa datalist)
    role_options = [
        "Head", "Deputy head", "Academic", "Academic 2", "Teacher",
        "Sport", "Sport 2", "Environment", "Environment 2", "Despline",
        "Matron", "Mnasihi(f)", "Mnasihi(m)", "Accountant", "Project", "Project 2",
        "Maintainance", "Maintainance 2", "Food"
    ]
    return render_template('edit_other_roles.html', teachers=teachers_data, roles=role_options)


#============================================= IPO CHINI =========================CHECK RESULTS=================
@app.route('/check_results', methods=['GET', 'POST'])
def check_results():
    # ------------------------------------------------------------
    # 1. Angalia kama kuna pin au phone kwenye URL (GET parameters)
    # ------------------------------------------------------------
    pin_param = request.args.get('pin', '').strip()
    phone_param = request.args.get('phone', '').strip()

    if pin_param or phone_param:
        student = None
        if phone_param:
            res = supabase.table('student_list').select('*').eq('parent_phone', phone_param).execute()
            if res.data:
                student = res.data[0]
        elif pin_param:
            try:
                sid = int(pin_param)
                res = supabase.table('student_list').select('*').eq('id', sid).execute()
                if res.data:
                    student = res.data[0]
            except ValueError:
                pass
        if student:
            # Pata matokeo kwa kutumia function iliyopo
            try:
                result = supabase.rpc('get_student_results_full', {'p_student_id': student['id']}).execute()
                data = result.data
                if data and 'error' in data:
                    return render_template('index.html', error=data['error'])
                if not data or 'student' not in data:
                    return render_template('index.html', error='Matokeo hayapatikani kwa mwanafunzi huyu.')
                return render_template('check_results.html', data=data)
            except Exception as e:
                return render_template('index.html', error=f'Hitilafu ya server: {str(e)}')
        else:
            return render_template('index.html', error='Hakuna mwanafunzi anayepatikana kwa taarifa ulizotoa.')

    # ------------------------------------------------------------
    # 2. Mantiki ya awali (POST / GET) – HAIJABADILISHWA
    # ------------------------------------------------------------
    if request.method == 'POST':
        phone = request.form.get('phone', '').strip()
        pin = request.form.get('pin', '').strip()

        student = None
        if phone:
            res = supabase.table('student_list').select('*').eq('parent_phone', phone).execute()
            if res.data:
                student = res.data[0]
        elif pin:
            try:
                sid = int(pin)
                res = supabase.table('student_list').select('*').eq('id', sid).execute()
                if res.data:
                    student = res.data[0]
            except ValueError:
                pass

        if not student:
            return render_template('index.html', error='Hakuna mwanafunzi anayepatikana kwa taarifa ulizotoa.')

        # Pata matokeo kwa kutumia function iliyopo
        try:
            result = supabase.rpc('get_student_results_full', {'p_student_id': student['id']}).execute()
            data = result.data
            if data and 'error' in data:
                return render_template('index.html', error=data['error'])
            if not data or 'student' not in data:
                return render_template('index.html', error='Matokeo hayapatikani kwa mwanafunzi huyu.')
            return render_template('check_results.html', data=data)
        except Exception as e:
            return render_template('index.html', error=f'Hitilafu ya server: {str(e)}')

    # GET request (hakuna parameters) – onyesha form ya kuingia
    return render_template('index.html')
#=================================================================================================check_results==

@app.route('/manage_subjects', methods=['GET'])
@login_required
def manage_subjects():
    # Orodha ya masomo
    subjects_res = supabase.table('subjects').select('id, subject_name').execute()
    subjects = subjects_res.data

    # Orodha ya madarasa (kidato)
    forms = ['Form1', 'Form2', 'Form3', 'Form4', 'Form5', 'Form6']

    # Chagua kidato kutoka query parameter (default = Form1)
    selected_form = request.args.get('form', 'Form1')
    # Chagua somo kutoka query parameter (default = somo la kwanza)
    selected_subject_id = request.args.get('subject_id', subjects[0]['id'] if subjects else None, type=int)

    # Pata wanafunzi wa kidato selected_form na walio ACTIVE
    students_res = supabase.table('student_list')\
        .select('id, student_name, current_class')\
        .eq('current_class', selected_form)\
        .eq('is_active', True)\
        .order('student_name')\
        .execute()
    students = students_res.data

    # Pata wanafunzi waliosajiliwa kwa somo hili (bila kujali kidato)
    registered_students = set()
    if selected_subject_id:
        reg_res = supabase.table('student_subjects')\
            .select('student_id')\
            .eq('subject_id', selected_subject_id)\
            .execute()
        registered_students = {r['student_id'] for r in reg_res.data}

    return render_template('manage_subjects.html',
                           subjects=subjects,
                           forms=forms,
                           selected_form=selected_form,
                           students=students,
                           selected_subject_id=selected_subject_id,
                           registered_students=registered_students)

@app.route('/save_subject_registrations', methods=['POST'])
@login_required
def save_subject_registrations():
    try:
        data = request.json
        subject_id = data.get('subject_id')
        student_ids = set(data.get('student_ids', []))
        form_class = data.get('form')   # 'Form1', 'Form2', nk.

        if not subject_id or not form_class:
            return jsonify({'error': 'Subject ID and form are required'}), 400

        # 1. Pata wanafunzi wote wa kidato hiki (ili kujua ni student_ids gani ni wa kidato hicho)
        students_in_class = supabase.table('student_list')\
            .select('id')\
            .eq('current_class', form_class)\
            .execute()
        class_student_ids = {s['id'] for s in students_in_class.data}

        # 2. Pata usajili uliopo kwa somo hili (kwa wanafunzi wa kidato hiki)
        existing_res = supabase.table('student_subjects')\
            .select('student_id')\
            .eq('subject_id', subject_id)\
            .execute()
        existing_all = {r['student_id'] for r in existing_res.data}
        # Chuja kwa wanafunzi wa kidato hiki
        existing_students = existing_all & class_student_ids

        # 3. Orodha ya wanafunzi waliochaguliwa (wanaostahili kidato hiki)
        selected_in_class = student_ids & class_student_ids

        to_add = selected_in_class - existing_students
        to_remove = existing_students - selected_in_class

        # 4. Ongeza wapya
        for sid in to_add:
            supabase.table('student_subjects').insert({
                'student_id': sid,
                'subject_id': subject_id
            }).execute()

        # 5. Futa wasiohitajika
        for sid in to_remove:
            supabase.table('student_subjects').delete()\
                .eq('subject_id', subject_id)\
                .eq('student_id', sid)\
                .execute()

        return jsonify({
            'message': f'Usajili umefanikiwa kwa {form_class}',
            'added': len(to_add),
            'removed': len(to_remove)
        })

    except Exception as e:
        # Kumbuka: Inapendekezwa kuandika error kwenye log (traceback.print_exc())
        return jsonify({'error': str(e)}), 500


@app.route('/results', methods=['GET', 'POST'])
@login_required   # Ongeza ikiwa unahitaji authentication
def results():
    if request.method == 'POST':
        class_name = request.form.get('form')
        exam_type_id = request.form.get('exam_type_id')
        year = request.form.get('year')
        term = request.form.get('term')

        if not all([class_name, exam_type_id, year, term]):
            return render_template('results.html', error='Tafadhali jaza vigezo vyote.')

        try:
            exam_type_id = int(exam_type_id)
            year = int(year)
            term = int(term)

            # Itisha Supabase function
            result = supabase.rpc('get_class_results', {
                'p_class': class_name,
                'p_exam_type_id': exam_type_id,
                'p_year': year,
                'p_term': term
            }).execute()

            students_data = result.data

            return render_template('results.html',
                                   students=students_data,
                                   selected_class=class_name,
                                   selected_exam_type=exam_type_id,
                                   selected_year=year,
                                   selected_term=term,
                                   exam_types=supabase.table('exam_types').select('id, exam_name').execute().data,
                                   forms=['Form1', 'Form2', 'Form3', 'Form4', 'Form5', 'Form6'],
                                   years=[2024, 2025, 2026],
                                   terms=[1, 2])
        except Exception as e:
            return render_template('results.html', error=str(e))

    # GET request: onyesha form tupu
    exam_types = supabase.table('exam_types').select('id, exam_name').execute().data
    print("Exam types data:", exam_types)
    return render_template('results.html',
                           students=[],
                           selected_class='Form 1',
                           selected_exam_type=exam_types[0]['id'] if exam_types else 1,
                           selected_year=2026,
                           selected_term=1,
                           exam_types=exam_types,
                           forms=['Form1', 'Form2', 'Form3', 'Form4', 'Form5', 'Form6'],
                           years=[2024, 2025, 2026],
                           terms=[1, 2])

#========= Activate WATORO =========================================================

@app.route('/manage_active', methods=['GET', 'POST'])
def manage_active():
    # ------------------------------------------------------------
    # AUTHENTICATION KUPITIA SESSION
    if 'user_id' not in session:
        return render_template('index.html', error='Hujaoana. Tafadhali ingia kwanza.')
    # ------------------------------------------------------------

    # Orodha ya madarasa
    forms = ['Form1', 'Form2', 'Form3', 'Form4', 'Form5', 'Form6']
    selected_form = request.args.get('form', 'Form1')

    if request.method == 'POST':
        try:
            data = request.json
            updates = data.get('updates', [])  # list ya {student_id, is_active}
            if not updates:
                return jsonify({'status': 'error', 'message': 'Hakuna mabadiliko'}), 400

            # Itisha Supabase function ya bulk update
            result = supabase.rpc('update_students_active', {'p_updates': updates}).execute()
            return jsonify({'status': 'success', 'message': 'Hali ya wanafunzi imesasishwa.'})
        except Exception as e:
            traceback.print_exc()
            return jsonify({'status': 'error', 'message': str(e)}), 500

    # GET: Onyesha wanafunzi wa form iliyochaguliwa
    try:
        result = supabase.rpc('get_students_by_form', {'p_form': selected_form}).execute()
        students = result.data
    except Exception as e:
        students = []
        print("Error fetching students:", e)

    return render_template('manage_active.html',
                           forms=forms,
                           selected_form=selected_form,
                           students=students)


@app.route('/teacher-on-duty')
def teacher_on_duty():
    return render_template('tod_form.html')

@app.route('/generate_tod_report', methods=['POST'])
def generate_tod_report():
    try:
        data = request.json
        attendance = data.get('attendance', {})

        # ----- HESABU JUMLA YA WALIOSAJILIWA (total_registered_all) -----
        total_registered_all = 0
        for form_key, form_data in attendance.items():
            for stream in form_data.get('streams', []):
                total_registered_all += stream.get('registered_boys', 0)
                total_registered_all += stream.get('registered_girls', 0)
        # ----------------------------------------------------------------
        # Prepare data for PDF
        report_data = {
            'tod_name': data.get('tod_name', ''),
            'date': data.get('date', ''),
            'day_of_week': data.get('day_of_week', ''),
            'food_cooked': data.get('food_cooked', ''),
            'food_condition': data.get('food_condition', ''),
            'food_bad_action': data.get('food_bad_action', ''),
            'ingredients_available': data.get('ingredients_available', ''),
            'missing_ingredients': data.get('missing_ingredients', ''),
            'cooking_on_time': data.get('cooking_on_time', ''),
            'cooking_delay_reason': data.get('cooking_delay_reason', ''),
            'no_food_reason': data.get('no_food_reason', ''),
            'area_clean': data.get('area_clean', ''),
            'food_covered': data.get('food_covered', ''),
            'served_on_time': data.get('served_on_time', ''),
            'food_enough': data.get('food_enough', ''),
            'not_enough_action': data.get('not_enough_action', ''),
            'student_miss': data.get('student_miss', ''),
            'miss_food_reason': data.get('miss_food_reason', ''),
            'teacher_serving': data.get('teacher_serving', ''),
            'teacher_quality': data.get('teacher_quality', ''),
            'morning_number': data.get('morning_number', ''),
            'cleanliness': data.get('cleanliness', ''),
            'other_activities': data.get('other_activities', ''),
            'events': data.get('events', ''),
            'security_report': data.get('security_report', ''),
            'challenges': data.get('challenges', ''),
            'suggestions': data.get('suggestions', ''),
            'tod_comments': data.get('tod_comments', ''),
            'evening_activities': data.get('evening_activities', ''),
            'attendance': data.get('attendance', {}),
            'total_registered_all': total_registered_all,   # <-- variable mpya
            'generation_date': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }

        html_content = render_template('tod_report.html', **report_data)
        pdf_file = HTML(string=html_content).write_pdf()

        response = send_file(
            io.BytesIO(pdf_file),
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f"tod_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        )
        return response
    except Exception as e:
        return jsonify({'error': str(e)}), 500



def call_zai_api_with_retry(payload, max_retries=3):
    """Tuma ombi kwa Z.ai API, ikishindwa kwa 429, jaribu tena hadi mara 3."""
    headers = {
        'Authorization': f'Bearer {ZAI_API_KEY}',
        'Content-Type': 'application/json'
    }
    for attempt in range(max_retries):
        try:
            response = requests.post(ZAI_API_URL, headers=headers, json=payload, timeout=60)
            if response.status_code == 200:
                return response.json()
            elif response.status_code == 429:
                # Kama ni kosa la rate limit, subiri kwa muda unaoongezeka (exponential backoff)
                wait_time = 2 ** attempt  # sekunde 1, 2, 4
                print(f"Attempt {attempt + 1}: Rate limit hit. Waiting {wait_time} seconds...")
                time.sleep(wait_time)
                continue
            else:
                # Kwa makosa mengine, rudisha ujumbe wa kosa
                return {"error": f"HTTP {response.status_code}: {response.text}"}
        except requests.exceptions.RequestException as e:
            if attempt == max_retries - 1:
                return {"error": f"Network error after {max_retries} attempts: {str(e)}"}
            time.sleep(2 ** attempt)
            continue
    return {"error": "Max retries exceeded"}

@app.route('/translate', methods=['POST'])
def translate_text():
    data = request.get_json()
    kiswahili = data.get('kiswahili', '').strip()
    if not kiswahili:
        return jsonify({'error': 'Nafasi tupu'}), 400

    payload = {
        'model': ZAI_MODEL,
        'messages': [
            {'role': 'system', 'content': SYSTEM_PROMPT},
            {'role': 'user', 'content': kiswahili}
        ],
        'temperature': 0.3,
        'max_tokens': 1000
    }

    result = call_zai_api_with_retry(payload)
    if 'error' in result:
        return jsonify({'error': result['error']}), 500

    try:
        translation = result['choices'][0]['message']['content'].strip()
        return jsonify({'translation': translation})
    except (KeyError, IndexError):
        return jsonify({'error': 'Invalid response structure from AI'}), 500


        #======= DUTY ROUTES  ====================================================



REFERENCE_DATE = datetime(2026, 4, 13).date()  # Jumatatu ya wiki ya kwanza

# ------------------------------------------------------------
# FUNGSI MSAIDIZI (SALAMA)
# ------------------------------------------------------------
def get_monday_of_week(date=None):
    if date is None:
        date = datetime.now().date()
    return date - timedelta(days=date.weekday())

def get_active_teachers():
    """Rudisha walimu wote wanaofanya kazi, wakiwa wamepangwa kwa rotation_order (NULLs last) kisha id."""
    try:
        result = supabase.table('teachers')\
            .select('id, name, rotation_order')\
            .eq('is_active', True)\
            .execute()
        teachers = result.data
        # Panga kwa rotation_order (NULLs last) kisha id (kwa mkono)
        teachers.sort(key=lambda x: (x.get('rotation_order') is None, x.get('rotation_order', 0), x['id']))
        return teachers
    except Exception as e:
        print("Error in get_active_teachers:", e)
        return []

def get_teacher_for_week(target_monday):
    teachers = get_active_teachers()
    if not teachers:
        return None
    weeks_diff = (target_monday - REFERENCE_DATE).days // 7
    if weeks_diff < 0:
        weeks_diff = 0
    index = weeks_diff % len(teachers)
    return teachers[index]

def get_current_duty():
    monday = get_monday_of_week()
    teacher = get_teacher_for_week(monday)
    if not teacher:
        return None
    week_end = monday + timedelta(days=6)
    return {
        'teacher': teacher,
        'week_start': monday.isoformat(),
        'week_end': week_end.isoformat()
    }

def get_next_duty_for_teacher(teacher_id):
    teachers = get_active_teachers()
    if not teachers:
        return None
    idx = None
    for i, t in enumerate(teachers):
        if t['id'] == teacher_id:
            idx = i
            break
    if idx is None:
        return None
    n = len(teachers)
    today_monday = get_monday_of_week()
    weeks_since_ref = (today_monday - REFERENCE_DATE).days // 7
    current_index = weeks_since_ref % n
    diff = (idx - current_index) % n
    if diff == 0:
        target_monday = today_monday
    else:
        target_monday = today_monday + timedelta(days=7 * diff)
    week_end = target_monday + timedelta(days=6)
    return {
        'start': target_monday.isoformat(),
        'end': week_end.isoformat()
    }

def get_duty_week_days(week_start):
    return [week_start + timedelta(days=i) for i in range(7)]

def get_report_status(teacher_id, date):
    try:
        result = supabase.table('duty_reports')\
            .select('id')\
            .eq('teacher_id', teacher_id)\
            .eq('report_date', date.isoformat())\
            .execute()
        return len(result.data) > 0
    except:
        return False

@app.route('/api/duty/all')
def api_duty_all():
    teachers = get_active_teachers()
    result = []
    for t in teachers:
        duty = get_next_duty_for_teacher(t['id'])
        if duty:
            result.append({
                'teacher_id': t['id'],
                'teacher_name': t['name'],
                'duty_start': duty['start'],
                'duty_end': duty['end']
            })
    return jsonify(result)

@app.route('/api/duty/current')
def api_duty_current():
    duty = get_current_duty()
    if not duty:
        return jsonify({'error': 'Hakuna walimu wanaofanya kazi'}), 404
    return jsonify(duty)

@app.route('/api/duty/current/week')
def api_duty_current_week():
    duty = get_current_duty()
    if not duty:
        return jsonify({'error': 'No duty teacher'}), 404
    week_start = datetime.fromisoformat(duty['week_start']).date()
    days = get_duty_week_days(week_start)
    teacher_id = duty['teacher']['id']
    week_data = []
    for day in days:
        has_report = get_report_status(teacher_id, day)
        week_data.append({
            'date': day.isoformat(),
            'has_report': has_report,
            'day_name': day.strftime('%A')
        })
    return jsonify({
        'teacher': duty['teacher'],
        'week_start': duty['week_start'],
        'week_end': duty['week_end'],
        'days': week_data
    })

@app.route('/duty-rotation')
def duty_rotation():
    return render_template('duty_rotation.html')


#===========SAVE TOD REPORT =======================================================


# ------------------------------------------------------------
# HELPER FUNCTIONS (for duty rotation logic)
# ------------------------------------------------------------

def get_teacher_for_date(target_date):
    """Kokotoa mwalimu anayepaswa kuwa zamu kwa tarehe yoyote."""
    target_monday = get_monday_of_week(target_date)
    # Rejesha matokeo kutoka kwa function iliyofanya kazi
    return get_teacher_for_week(target_monday)

def get_active_teachers():
    result = supabase.table('teachers')\
        .select('id, name, rotation_order, teachers_table_id')\
        .eq('is_active', True)\
        .execute()
    teachers = result.data
    teachers.sort(key=lambda x: (x.get('rotation_order') is None, x.get('rotation_order', 0), x['id']))
    return teachers

def upload_pdf_to_storage(teacher_id, report_date, pdf_bytes):
    """Pakilia PDF kwenye bucket, rudisha storage path."""
    filename = f"teacher_{teacher_id}_{report_date}.pdf"
    file_path = f"{report_date}/{filename}"
    supabase.storage.from_(STORAGE_BUCKET).upload(
        file_path, pdf_bytes, file_options={"content-type": "application/pdf"}
    )
    return file_path

def get_signed_pdf_url(storage_path, expiry=15768000):
    """Zalisha signed URL ya muda mfupi."""
    res = supabase.storage.from_(STORAGE_BUCKET).create_signed_url(storage_path, expiry)
    return res['signedURL']

# ------------------------------------------------------------
# ROUTE: HIFADHI RIPOTI
# ------------------------------------------------------------
@app.route('/save-tod-report', methods=['POST'])
def save_tod_report():
    data = request.json
    if not data:
        return jsonify({'error': 'No data received'}), 400

    # ------------------------------------------------------------
    # 1. ANGALIA USER ALIYEINGIA
    # ------------------------------------------------------------
    if 'user_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    user_id = session['user_id']

    # ------------------------------------------------------------
    # 2. CHUKUA TAREHE KUTOKA DATA (HII ILIKOSEKANA)
    # ------------------------------------------------------------
    date = data.get('date', '')
    if not date:
        return jsonify({'error': 'Date is missing'}), 400

    # ------------------------------------------------------------
    # 3. HESABU JUMLA YA WALIOSAJILIWA (total_registered_all)
    # ------------------------------------------------------------
    attendance = data.get('attendance', {})
    total_registered_all = 0
    for form_key, form_data in attendance.items():
        for stream in form_data.get('streams', []):
            total_registered_all += stream.get('registered_boys', 0)
            total_registered_all += stream.get('registered_girls', 0)

    # ------------------------------------------------------------
    # 4. PATA MWALIMU WA ZAMU KWA TAREHE HII
    # ------------------------------------------------------------
    try:
        report_date_obj = datetime.strptime(date, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'error': f'Invalid date format: "{date}". Expected YYYY-MM-DD.'}), 400

    teacher_on_duty = get_teacher_for_date(report_date_obj)
    if not teacher_on_duty:
        return jsonify({'error': 'No active teacher for this date'}), 404
    teacher_id = teacher_on_duty['id']

    # ------------------------------------------------------------
    # 5. TENGENEZA PDF KWA KUTUMIA TEMPLATE
    # ------------------------------------------------------------
    try:
        report_data = {
            'tod_name': data.get('tod_name', ''),
            'date': date,
            'day_of_week': data.get('day_of_week', ''),
            'morning_number': data.get('morning_number', ''),
            'cleanliness': data.get('cleanliness', ''),
            'other_activities': data.get('other_activities', ''),
            'food_cooked': data.get('food_cooked', ''),
            'food_condition': data.get('food_condition', ''),
            'food_bad_action': data.get('food_bad_action', ''),
            'ingredients_available': data.get('ingredients_available', ''),
            'missing_ingredients': data.get('missing_ingredients', ''),
            'cooking_on_time': data.get('cooking_on_time', ''),
            'cooking_delay_reason': data.get('cooking_delay_reason', ''),
            'no_food_reason': data.get('no_food_reason', ''),
            'area_clean': data.get('area_clean', ''),
            'food_covered': data.get('food_covered', ''),
            'served_on_time': data.get('served_on_time', ''),
            'food_enough': data.get('food_enough', ''),
            'not_enough_action': data.get('not_enough_action', ''),
            'student_miss': data.get('student_miss', ''),
            'miss_food_reason': data.get('miss_food_reason', ''),
            'teacher_serving': data.get('teacher_serving', ''),
            'teacher_quality': data.get('teacher_quality', ''),
            'events': data.get('events', ''),
            'security_report': data.get('security_report', ''),
            'challenges': data.get('challenges', ''),
            'suggestions': data.get('suggestions', ''),
            'tod_comments': data.get('tod_comments', ''),
            'evening_activities': data.get('evening_activities', ''),
            'attendance': attendance,
            'total_registered_all': total_registered_all,
            'generation_date': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        html_content = render_template('tod_report.html', **report_data)
        pdf_bytes = HTML(string=html_content).write_pdf()
    except Exception as e:
        return jsonify({'error': f'PDF generation failed: {str(e)}'}), 500

    # ------------------------------------------------------------
    # 6. PAKILIA PDF KWENYE BUCKET
    # ------------------------------------------------------------
    try:
        storage_path = upload_pdf_to_storage(teacher_id, date, pdf_bytes)
    except Exception as e:
        return jsonify({'error': f'Upload to storage failed: {str(e)}'}), 500

    # ------------------------------------------------------------
    # 7. HIFADHI REKODI KWENYE DUTY_REPORTS
    # ------------------------------------------------------------
    try:
        supabase.table('duty_reports').insert({
            'teacher_id': teacher_id,
            'user_id': user_id,
            'report_date': date,
            'storage_path': storage_path
        }).execute()
    except Exception as e:
        return jsonify({'error': f'Database error: {str(e)}'}), 500

    # ------------------------------------------------------------
    # 8. ONYESHA PDF (VIEW)
    # ------------------------------------------------------------
    signed_url = get_signed_pdf_url(storage_path, expiry=15768000)
    return render_template('view_pdf.html', pdf_url=signed_url, report_date=date)

#=================================== view report from starage ==============================================
@app.route('/view-report')
def view_report():
    # 1. Angalia kama mtumiaji ameingia
    if 'user_id' not in session:
        return "Unauthorized. Please login.", 401

    # 2. Pata tarehe kutoka query parameter (?date=YYYY-MM-DD)
    report_date = request.args.get('date')
    if not report_date:
        return "Missing date parameter", 400

    # 3. Thibitisha umbizo la tarehe
    try:
        date_obj = datetime.strptime(report_date, '%Y-%m-%d').date()
    except ValueError:
        return "Invalid date format. Use YYYY-MM-DD", 400

    # 4. Pata mwalimu anayepaswa kuwa zamu kwa tarehe hii (tumia function uliyonayo)
    teacher_on_duty = get_teacher_for_date(date_obj)  # tumia function iliyopo
    if not teacher_on_duty:
        return "No teacher found for this date", 404
    teacher_id = teacher_on_duty['id']

    # 5. Tafuta storage_path kwenye duty_reports
    result = supabase.table('duty_reports')\
        .select('storage_path')\
        .eq('teacher_id', teacher_id)\
        .eq('report_date', report_date)\
        .execute()
    if not result.data:
        return "No report found for this date", 404

    storage_path = result.data[0]['storage_path']

    # 6. Zalisha signed URL (miezi 6)
    try:
        signed_url = get_signed_pdf_url(storage_path, expiry=15768000)  # 6 months
    except Exception as e:
        return f"Error generating PDF URL: {e}", 500

    # 7. Onyesha ukurasa wa ku-view PDF (tumia template `view_pdf.html`)
    return render_template('view_pdf.html', pdf_url=signed_url, report_date=report_date)


@app.route('/download-duty-roster')
def download_duty_roster():
    # Pata data ya walimu wote na zamu zao (kama ilivyo kwa /api/duty/all)
    teachers = get_active_teachers()
    roster_data = []
    for t in teachers:
        duty = get_next_duty_for_teacher(t['id'])
        if duty:
            roster_data.append({
                'name': t['name'],
                'start': duty['start'],
                'end': duty['end']
            })

    # Panga kwa tarehe ya kuanza (ili kudumisha utaratibu)
    roster_data.sort(key=lambda x: x['start'])

    # Ongeza namba za serial (1-indexed)
    for idx, item in enumerate(roster_data, start=1):
        item['serial'] = idx

    # Tafuta mwaka (kutoka tarehe ya kwanza au ya sasa)
    if roster_data:
        first_start = roster_data[0]['start']
        year = first_start[:4] if first_start else str(datetime.now().year)
    else:
        year = str(datetime.now().year)

    # Tengeneza PDF kwa kutumia template
    html_content = render_template('duty_roster_pdf.html',
                                   roster=roster_data,
                                   year=year,
                                   generated_date=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    pdf = HTML(string=html_content).write_pdf()

    # Rudisha PDF kama attachment (ili ipakuliwe)
    return send_file(
        io.BytesIO(pdf),
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'duty_roster_{year}.pdf'
    )


@app.route('/duty-history')
def duty_history():
    teacher_id = request.args.get('teacher_id', 'all')
    try:
        if teacher_id == 'all':
            result = supabase.rpc('get_duty_history_by_teacher', {'p_teacher_id': None}).execute()
        else:
            teacher_id_int = int(teacher_id)
            result = supabase.rpc('get_duty_history_by_teacher', {'p_teacher_id': teacher_id_int}).execute()
        reports = result.data
    except Exception as e:
        reports = []
        error = str(e)
        return render_template('duty_history.html', reports=reports, error=error, teachers=[], selected_teacher='all')

    # Pata orodha ya walimu wote (kwa dropdown)
    try:
        teachers_res = supabase.table('teachers')\
            .select('id, name')\
            .eq('is_active', True)\
            .order('name')\
            .execute()
        teachers = teachers_res.data
    except:
        teachers = []

    return render_template('duty_history.html',
                           reports=reports,
                           teachers=teachers,
                           selected_teacher=teacher_id)


@app.route('/api/update-week-status', methods=['POST'])
def update_week_status():
    # 1. Angalia kama mtumiaji ameingia (hiari)
    if 'user_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401   # Hapa ndipo palikuwa na kosa

    # 2. Pata mwalimu wa zamu wa wiki hii
    current_duty = get_current_duty()
    if not current_duty:
        return jsonify({'error': 'No teacher on duty for this week'}), 404

    duty_teacher = current_duty['teacher']
    teachers_table_id = duty_teacher.get('teachers_table_id')

    if not teachers_table_id:
        return jsonify({'error': f'Teacher {duty_teacher["name"]} has no teachers_table_id link'}), 404

    # 3. Sasisha week_status
    try:
        # Weka 'off' kwa walimu wote (au kwa wale active tu)
        supabase.table('teachers_table')\
            .update({'week_status': 'off'})\
            .neq('id', 0)\
            .execute()

        # Weka 'on_duty' kwa mwalimu wa zamu
        supabase.table('teachers_table')\
            .update({'week_status': 'on_duty'})\
            .eq('id', teachers_table_id)\
            .execute()

        return jsonify({'message': f"Week status updated. Teacher {duty_teacher['name']} is now on duty."})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
@app.route('/sitting-plan')
def sitting_plan():
    return render_template('sitting_plan.html')
@app.route('/exam-table')
def exam_table():
    return render_template('exam_table.html')

#===================================================  Exam Types  ==========================
@app.route('/manage-exam-types', methods=['GET', 'POST'])
def manage_exam_types():
    # Angalia kama mtumiaji ameingia (hiari, unaweza kuongeza role check)
    if 'user_id' not in session:
        return redirect(url_for('login'))

    if request.method == 'POST':
        # Pata data ya JSON yenye mapendekezo: [{'id': 1, 'is_done': True, 'is_allowed_view': False}, ...]
        updates = request.json.get('updates', [])
        if not updates:
            return jsonify({'error': 'Hakuna mabadiliko'}), 400

        try:
            # Fanya update kwa kila rekodi (unaweza kutumia loop au upsert)
            for item in updates:
                supabase.table('exam_types')\
                    .update({
                        'is_done': item['is_done'],
                        'is_allowed_view': item['is_allowed_view']
                    })\
                    .eq('id', item['id'])\
                    .execute()
            return jsonify({'message': 'Hali zimesasishwa kikamilifu.'})
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    # GET: onyesha orodha ya exam_types
    try:
        result = supabase.table('exam_types').select('*').order('id').execute()
        exam_types = result.data
        return render_template('manage_exam_types.html', exam_types=exam_types)
    except Exception as e:
        return f"Error fetching data: {e}", 500

@app.route('/add-exam-type', methods=['POST'])
def add_exam_type():
    # Angalia kama mtumiaji ameingia (hiari)
    if 'user_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401

    data = request.json
    exam_name = data.get('exam_name', '').strip()
    if not exam_name:
        return jsonify({'error': 'Jina la mtihani linahitajika'}), 400

    try:
        # Ingiza mtihani mpya (default is_done=false, is_allowed_view=false)
        result = supabase.table('exam_types').insert({
            'exam_name': exam_name,
            'is_done': False,
            'is_allowed_view': False
        }).execute()
        return jsonify({'message': 'Mtihani umeongezwa kikamilifu', 'new_id': result.data[0]['id']})
    except Exception as e:
        return jsonify({'error': str(e)}), 500



#================================================== BARUA YA KIKAZI ====================
@app.route('/writing-latter')
def writing_latter():
    return render_template('latter.html')

#=============================================== ADD NEW STUDENTS =================

from flask import render_template, request, redirect, url_for, flash, jsonify
from datetime import datetime
import pandas as pd
import io

# ------------------------------
# 1. Ongezeko la mwanafunzi mmoja (tayari ulikuwa nayo)
# ------------------------------
@app.route('/add-student', methods=['GET', 'POST'])
def add_student():
    if request.method == 'POST':
        student_name = request.form.get('student_name', '').strip()
        sex = request.form.get('sex', '')
        current_class = request.form.get('current_class', '')
        parent_phone = request.form.get('parent_phone', '').strip()
        stream = request.form.get('stream', '').strip()
        admission_year = request.form.get('admission_year', type=int)
        is_active = True if request.form.get('is_active') else False

        # Sehemu za lazima: jina, jinsia, kidato
        if not student_name or not sex or not current_class:
            flash('Tafadhali jaza sehemu: Jina la mwanafunzi, Jinsia, na Kidato.', 'error')
            return redirect(url_for('add_student'))

        # Ikiwa parent_phone ni tupu, weka None (sio kamba tupu)
        if not parent_phone:
            parent_phone = None

        if admission_year is None:
            admission_year = datetime.now().year

        try:
            supabase.table('student_list').insert({
                'student_name': student_name,
                'sex': sex,
                'current_class': current_class,
                'parent_phone': parent_phone,
                'stream': stream if stream else None,  # weka None ikiwa tupu
                'admission_year': admission_year,
                'is_active': is_active
            }).execute()
            flash(f'Mwanafunzi {student_name} ameongezwa kikamilifu!', 'success')
            return redirect(url_for('add_student'))
        except Exception as e:
            flash(f'Kosa la database: {str(e)}', 'error')
            return redirect(url_for('add_student'))

    current_year = datetime.now().year
    streams = ['A', 'B', 'C', 'D', 'E', 'F']
    return render_template('add_student.html', current_year=current_year, streams=streams)
# ------------------------------
# 2. API ya kupata mwanafunzi wa mwisho (id kubwa)
# ------------------------------
@app.route('/api/latest-student')
def api_latest_student():
    try:
        response = supabase.rpc('get_latest_student').execute()
        data = response.data
        if data and 'error' in data:
            return jsonify({'error': data['error']}), 404
        return jsonify(data)
    except Exception as e:
        print("RPC error:", e)
        return jsonify({'error': 'Muunganisho umeshindwa. Jaribu tena.'}), 500


# ------------------------------
# 3. Upload Excel na kuongeza wanafunzi kwa wingi
# ------------------------------
@app.route('/upload-students-excel', methods=['POST'])
def upload_students_excel():
    if 'excel_file' not in request.files:
        flash('Hakuna faili iliyochaguliwa.', 'error')
        return redirect(url_for('add_student'))

    file = request.files['excel_file']
    if file.filename == '':
        flash('Tafadhali chagua faili.', 'error')
        return redirect(url_for('add_student'))

    if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        flash('Faili lazima iwe ya aina ya Excel (.xlsx au .xls).', 'error')
        return redirect(url_for('add_student'))

    try:
        excel_data = pd.ExcelFile(file)
        total_added = 0
        errors = []

        for sheet_name in excel_data.sheet_names:
            current_class = sheet_name.strip()
            if not current_class.startswith('Form'):
                errors.append(f"Sheet '{sheet_name}' ina jina lisilofaa (lazima liwe kama 'Form4'). Imepuuzwa.")
                continue

            df = pd.read_excel(file, sheet_name=sheet_name)
            if 'student_name' not in df.columns or 'sex' not in df.columns:
                errors.append(f"Sheet '{sheet_name}' haina safu muhimu (student_name, sex). Imepuuzwa.")
                continue

            # Badilisha NaN na 'nan' (kamba) kuwa None
            df = df.replace([pd.NA, 'nan', 'NaN', 'NAN'], None)
            df = df.where(pd.notnull(df), None)

            for idx, row in df.iterrows():
                student_name = row.get('student_name')
                sex = row.get('sex')
                parent_phone = row.get('parent_phone')
                stream = row.get('stream')
                admission_year = row.get('admission_year')
                is_active_val = row.get('is_active')

                if not student_name or not sex:
                    errors.append(f"Sheet '{sheet_name}': Safu {idx+2} haina jina au jinsia.")
                    continue
                if sex not in ['M', 'F']:
                    errors.append(f"Sheet '{sheet_name}': Jinsia '{sex}' si sahihi.")
                    continue

                # parent_phone
                if parent_phone is None or str(parent_phone).lower() == 'nan':
                    parent_phone = None
                else:
                    parent_phone = str(parent_phone).strip()
                    if parent_phone == '':
                        parent_phone = None

                # stream
                if stream is None or str(stream).lower() == 'nan':
                    stream = None
                else:
                    stream = str(stream).strip()
                    if stream == '':
                        stream = None

                # admission_year: kama ni NaN, None, au si namba, tumia 2026 (au badilisha)
                if admission_year is None or str(admission_year).lower() == 'nan':
                    admission_year = 2026
                else:
                    try:
                        # Geuza kuwa integer (kwa kuwa Excel wakati mwingine huleta float)
                        admission_year = int(float(admission_year))
                        # Ikiwa namba ni ya kushangaza (k.m. 1,2,3 kutokana na drag), rejesha 2026
                        if admission_year < 2000 or admission_year > 2030:
                            admission_year = 2026
                    except:
                        admission_year = 2026

                # is_active
                if is_active_val is None or str(is_active_val).lower() == 'nan':
                    is_active = True
                else:
                    if isinstance(is_active_val, bool):
                        is_active = is_active_val
                    elif str(is_active_val).lower() in ['true', '1', 'yes', 't']:
                        is_active = True
                    else:
                        is_active = False

                try:
                    supabase.table('student_list').insert({
                        'student_name': student_name,
                        'sex': sex,
                        'current_class': current_class,
                        'parent_phone': parent_phone,
                        'stream': stream,
                        'admission_year': admission_year,
                        'is_active': is_active
                    }).execute()
                    total_added += 1
                except Exception as e:
                    errors.append(f"Sheet '{sheet_name}': Kushindwa kuongeza {student_name} - {str(e)}")

        flash_msg = f"{total_added} wanafunzi wameongezwa kikamilifu."
        if errors:
            flash_msg += f" Makosa: {'; '.join(errors[:5])}" + ('...' if len(errors)>5 else '')
        flash(flash_msg, 'success' if total_added>0 else 'error')

    except Exception as e:
        flash(f'Kosa la kusoma faili: {str(e)}', 'error')

    return redirect(url_for('add_student'))

@app.route('/download-excel-template')
def download_excel_template():
    # Unda data frame yenye headers na safu moja ya mfano (hiari)
    template_data = {
        'student_name': ['Mfano: Aisha Juma'],
        'sex': ['F'],
        'parent_phone': ['0712345678'],
        'stream': ['A'],
        'admission_year': [2026],
        'is_active': [True]
    }
    df = pd.DataFrame(template_data)

    # Unda faili la Excel kwenye memory
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Kwa kila kidato, unaweza kuweka sheet tofauti; hapa tunaweka sheet moja ya mfano
        df.to_excel(writer, sheet_name='Form4', index=False)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='student_upload_template.xlsx'
    )
#==================================================================================
#                        UPDATE STREAM  &  PARENT_PHONE
#==================================================================================

#==================================================================================
#                        UPDATE STREAM  &  PARENT_PHONE
#==================================================================================

@app.route('/edit-students', methods=['GET', 'POST'])
def edit_students():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    if request.method == 'POST':
        updates = request.json.get('updates', [])
        if not updates:
            return jsonify({'error': 'Hakuna mabadiliko'}), 400
        try:
            for item in updates:
                supabase.table('student_list')\
                    .update({
                        'stream': item.get('stream'),
                        'parent_phone': item.get('parent_phone')
                    })\
                    .eq('id', item['id'])\
                    .execute()
            return jsonify({'message': 'Mabadiliko yamehifadhiwa.'})
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    # GET: Retrieve students
    selected_class = request.args.get('class', '')
    query = supabase.table('student_list').select('id, student_name, current_class, stream, parent_phone, sex')
    if selected_class:
        query = query.eq('current_class', selected_class)
    result = query.order('id').execute()
    students = result.data

    # Ensure every student has an 'id' (Supabase should return 'id' if primary key is 'id')
    # If your primary key is named differently (e.g., 'student_id'), rename it:
    for s in students:
        if 'id' not in s and 'student_id' in s:
            s['id'] = s['student_id']

    classes = ['Form1', 'Form2', 'Form3', 'Form4', 'Form5', 'Form6']
    return render_template('edit_students.html', students=students, classes=classes, selected_class=selected_class)


# ------------------------------------------------------------
#             SHARING FILES
# ------------------------------------------------------------

STORAGE_BUCKET1 = "shared_files"
ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'docx', 'pdf', 'csv'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# ------------------------------------------------------------
# ROUTE YA KUPATA ORODHA YA WATUMIAJI (kwa dropdown)
# ------------------------------------------------------------
@app.route('/api/users-list')
def api_users_list():
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    result = supabase.table('users_table').select('username, role').execute()
    # Orodha ya usernames, isipokuwa mtumiaji mwenyewe
    current_user = session['username']
    users = [{'username': u['username'], 'role': u['role']} for u in result.data]
    return jsonify(users)

# ------------------------------------------------------------
# ROUTE YA KUTUMA FAILI (FORM)
# ------------------------------------------------------------
@app.route('/send-file', methods=['GET', 'POST'])
def send_file_page():
    if 'username' not in session:
        return redirect(url_for('login'))

    if request.method == 'POST':
        sender = session['username']
        recipient = request.form.get('recipient')
        use_for = request.form.get('use_for', '').strip()
        uploaded_file = request.files.get('file')

        if not recipient or not use_for or not uploaded_file:
            flash('Tafadhali jaza sehemu zote na chagua faili.', 'error')
            return redirect(url_for('send_file_page'))

        if not allowed_file(uploaded_file.filename):
            flash('Aina ya faili hairuhusiwi. Ruhusiwa: xlsx, xls, docx, pdf, csv', 'error')
            return redirect(url_for('send_file_page'))

        filename = secure_filename(uploaded_file.filename)
        # Jina la faili liwe na timestamp ili kuepuka duplicate
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        unique_filename = f"{timestamp}_{filename}"
        file_path = f"{sender}/{unique_filename}"

        try:
            # Pakia faili kwenye bucket
            file_bytes = uploaded_file.read()
            supabase.storage.from_(STORAGE_BUCKET1).upload(file_path, file_bytes, file_options={"content-type": uploaded_file.content_type})
            # Pata signed URL (miezi 3 halali)
            signed_url = supabase.storage.from_(STORAGE_BUCKET1).create_signed_url(file_path, 7776000)  # 90 days
            file_url = signed_url['signedURL']

            # Ingiza rekodi kwenye shared_files
            supabase.table('shared_files').insert({
                'sender': sender,
                'recipient': recipient,
                'use_for': use_for,
                'file_url': file_url,
                'sent_at': datetime.now().isoformat()
            }).execute()

            flash('Faili limetumwa kikamilifu!', 'success')
            return redirect(url_for('send_file_page'))
        except Exception as e:
            flash(f'Kosa la kuwasilisha faili: {str(e)}', 'error')
            return redirect(url_for('send_file_page'))

    # GET: onyesha form
    return render_template('send_file.html')

# ------------------------------------------------------------
# ROUTE YA KUONA FAILI ZANGU (AJAX)
# ------------------------------------------------------------
@app.route('/my-files')
def my_files():
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401

    current_user = session['username']
    current_role = session.get('role', '')
    three_months_ago = datetime.now() - timedelta(days=90)

    # Pata faili zote za muda huu (sio nyingi kwa kawaida)
    result = supabase.table('shared_files')\
        .select('*')\
        .gte('sent_at', three_months_ago.isoformat())\
        .execute()
    all_files = result.data

    if current_role == 'Academic':
        # Pata orodha ya majina ya watumiaji academic (herufi ndogo kwa ulinganifu)
        users_res = supabase.table('users_table').select('username').eq('role', 'Academic').execute()
        academic_usernames = [u['username'].lower() for u in users_res.data]

        # Chuja faili: mtumaji ni mwenyewe, au mpokeaji yuko kwenye orodha ya academic
        files = []
        for f in all_files:
            sender_lower = f['sender'].lower()
            recipient_lower = f['recipient'].lower()
            if sender_lower == current_user.lower() or recipient_lower in academic_usernames:
                files.append(f)
    else:
        # Watumiaji wa kawaida: wanaona faili walizotuma au walizopokea
        files = [f for f in all_files if f['sender'] == current_user or f['recipient'] == current_user]

    # Panga kwa sent_at descending
    files.sort(key=lambda x: x['sent_at'], reverse=True)
    return jsonify(files)
# ------------------------------------------------------------
# ROUTE YA KUBADILISHA JINA LA FAILI (DOWNLOAD) – HIARI
# ------------------------------------------------------------
@app.route('/download-shared-file')
def download_shared_file():
    # Pata file_url kutoka query parameter
    file_url = request.args.get('url')
    if not file_url:
        return "Missing URL", 400
    # Tuma mkasi wa kuwa download (kwa kuwa signed URL inaweza kutumika moja kwa moja, unaweza kuelekeza)
    # Lakini kwa urahisi, rudisha redirect kwa signed URL (itapakuliwa)
    return redirect(file_url)
#============================================================================
#                                 ANALYSIS YA MTIHANI MZIMA
#=============================================================================
#============================================================================
#                                 ANALYSIS YA MTIHANI MZIMA
#=============================================================================
@app.route('/analysis')
def analysis():
    # Pata orodha ya aina za mitihani ambazo zina data (zilizofanywa)
    # Tunachagua exam_types ambapo is_done = true (kwa mfano)
    # Au tunaweza kuchukua kutoka exam_registration
    exam_types_res = supabase.table('exam_types')\
        .select('id, exam_name')\
        .eq('is_done', True)\
        .execute()
    exam_types = exam_types_res.data
    # Ikiwa hakuna, angalia alternative: chukua distinct exam_type_id kutoka exam_registration
    if not exam_types:
        # Chukua kutoka exam_registration
        distinct_res = supabase.table('exam_registration')\
            .select('exam_type_id, exam_types(exam_name)')\
            .execute()
        # Kukusanya unique
        seen = set()
        exam_types = []
        for item in distinct_res.data:
            eid = item['exam_type_id']
            if eid not in seen:
                seen.add(eid)
                exam_types.append({'id': eid, 'exam_name': item['exam_types']['exam_name']})

    # Orodha ya madarasa
    classes = ['Form1', 'Form2', 'Form3', 'Form4', 'Form5', 'Form6']
    return render_template('analysis.html', exam_types=exam_types, classes=classes)

@app.route('/api/analysis-data', methods=['POST'])
def analysis_data():
    data = request.json
    exam_type_id = data.get('exam_type_id')
    class_name = data.get('class')
    if not exam_type_id or not class_name:
        return jsonify({'error': 'Missing parameters'}), 400
    try:
        result = supabase.rpc('get_exam_analysis', {
            'p_exam_type_id': int(exam_type_id),
            'p_class': class_name
        }).execute()
        return jsonify(result.data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/manage_duty_teachers', methods=['GET', 'POST'])
@login_required
def manage_duty_teachers():
    # Ruhusa: Academic au Admin pekee
    if session.get('role') not in ['Academic', 'Admin']:
        flash("Huna ruhusa ya kufikia ukurasa huu.", "error")
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        teachers_table_id = request.form.get('teachers_table_id', type=int)
        is_active = True if request.form.get('is_active') else False

        if not name or not teachers_table_id:
            flash("Tafadhali jaza jina la mwalimu na uchague mwalimu kutoka orodha.", "error")
            return redirect(url_for('manage_duty_teachers'))

        # Thibitisha kuwa mwalimu huyo (kwa teachers_table_id) hajasajiliwa tayari
        existing = supabase.table('teachers').select('id').eq('teachers_table_id', teachers_table_id).execute()
        if existing.data:
            flash("Mwalimu huyo tayari amesajiliwa kwenye orodha ya zamu.", "error")
            return redirect(url_for('manage_duty_teachers'))

        # Pata rotation_order inayofuata
        max_order_res = supabase.table('teachers').select('rotation_order').order('rotation_order', desc=True).limit(1).execute()
        next_order = (max_order_res.data[0]['rotation_order'] + 1) if max_order_res.data else 1

        try:
            supabase.table('teachers').insert({
                'name': name,
                'teachers_table_id': teachers_table_id,
                'rotation_order': next_order,
                'is_active': is_active
            }).execute()
            flash(f"Mwalimu '{name}' ameongezwa kikamilifu.", "success")
        except Exception as e:
            flash(f"Hitilafu: {str(e)}", "error")
        return redirect(url_for('manage_duty_teachers'))

    # GET: Onyesha fomu na orodha ya walimu wa zamu
    # Pata orodha ya walimu kutoka teachers_table (kwa dropdown)
    teachers_table_data = supabase.table('teachers_table').select('id, teacher_name').order('teacher_name').execute()
    available_teachers = teachers_table_data.data

    # Pata walimu wote wa zamu (pamoja na join kupata jina kutoka teachers_table)
    duty_teachers_res = supabase.table('teachers')\
        .select('*, teachers_table(teacher_name)')\
        .order('rotation_order')\
        .execute()
    duty_teachers = []
    for t in duty_teachers_res.data:
        duty_teachers.append({
            'id': t['id'],
            'name': t['name'],
            'rotation_order': t['rotation_order'],
            'is_active': t['is_active'],
            'teachers_table_id': t['teachers_table_id'],
            'teacher_name_from_table': t['teachers_table']['teacher_name'] if t['teachers_table'] else 'N/A'
        })

    return render_template('manage_duty_teachers.html',
                           available_teachers=available_teachers,
                           duty_teachers=duty_teachers)

@app.route('/manage_duty_teachers/toggle_active', methods=['POST'])
@login_required
def toggle_duty_teacher_active():
    data = request.json
    teacher_id = data.get('id')
    is_active = data.get('is_active')
    if teacher_id is None or is_active is None:
        return jsonify({'error': 'Missing parameters'}), 400
    try:
        supabase.table('teachers').update({'is_active': is_active}).eq('id', teacher_id).execute()
        return jsonify({'message': 'Status updated'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/manage_duty_teachers/delete', methods=['POST'])
@login_required
def delete_duty_teacher():
    data = request.json
    teacher_id = data.get('id')
    if not teacher_id:
        return jsonify({'error': 'ID required'}), 400
    try:
        supabase.table('teachers').delete().eq('id', teacher_id).execute()
        return jsonify({'message': 'Teacher removed from duty rotation'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ========== Route ya kupata mwanafunzi mmoja (si lazima, tumia moja kwa moja kwenye update) ==========
# Haitumiki moja kwa moja, lakini ikiwa unahitaji.

# ========== Route ya kuhariri mwanafunzi mmoja (popup) ==========
@app.route('/update-student/<int:id>', methods=['POST'])
def update_student(id):
    if 'user_id' not in session:
        return jsonify({'error': 'Haujaingia'}), 401

    data = request.json
    student_name = data.get('student_name')
    gender = data.get('gender')
    current_class = data.get('current_class')

    if not student_name or not current_class:
        return jsonify({'error': 'Jina na kidato vinahitajika'}), 400

    try:
        supabase.table('student_list')\
            .update({
                'student_name': student_name,
                'sex': gender if gender else None,
                'current_class': current_class
            })\
            .eq('id', id)\
            .execute()
        return jsonify({'message': 'Taarifa zimehifadhiwa'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ========== Route ya kufuta mwanafunzi (pamoja na relations zake) ==========
@app.route('/delete-student/<int:id>', methods=['POST'])
def delete_student(id):
    if 'user_id' not in session:
        return redirect(url_for('login'))

    try:
        # Hatua 1: Pata exam_registration IDs zote za mwanafunzi huyu
        exam_regs = supabase.table('exam_registration')\
            .select('id')\
            .eq('student_id', id)\
            .execute()

        exam_reg_ids = [reg['id'] for reg in exam_regs.data]

        # Hatua 2: Futa marks_table kwa exam_registration_ids
        if exam_reg_ids:
            supabase.table('marks_table')\
                .delete()\
                .in_('exam_registration_id', exam_reg_ids)\
                .execute()

        # Hatua 3: Futa exam_registration kwa student_id
        supabase.table('exam_registration')\
            .delete()\
            .eq('student_id', id)\
            .execute()

        # Hatua 4: Futa student_subjects kama ipo (ikitumia student_id)
        # Angalia kama jedwali la student_subjects lipo na lina student_id
        # Kama halipo, ruka.
        try:
            supabase.table('student_subjects')\
                .delete()\
                .eq('student_id', id)\
                .execute()
        except:
            pass  # Jedwali halipo au haina column hiyo

        # Hatua 5: Futa mwanafunzi mwenyewe
        supabase.table('student_list')\
            .delete()\
            .eq('id', id)\
            .execute()

        flash('Mwanafunzi na data zake zimefutwa kabisa', 'success')
    except Exception as e:
        flash(f'Hitilafu wakati wa kufuta: {str(e)}', 'danger')

    return redirect(url_for('edit_students'))

# ========== Route yako ya awali ya edit_students (batch edit stream & phone) haibadilishwi ==========
# Iko kama ulivyoitoa, inafanya kazi vizuri.

# ======================= HARIRI JINA LA MTIHANI =======================
@app.route('/edit-exam-type', methods=['POST'])
def edit_exam_type():
    if 'user_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401

    data = request.json
    exam_id = data.get('id')
    new_name = data.get('exam_name', '').strip()

    if not exam_id or not new_name:
        return jsonify({'error': 'Jina na ID vinahitajika'}), 400

    try:
        supabase.table('exam_types')\
            .update({'exam_name': new_name})\
            .eq('id', exam_id)\
            .execute()
        return jsonify({'message': 'Jina la mtihani limebadilishwa kikamilifu.'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ======================= FUTA MTIHANI PAMOJA NA RELATION ZAKE =======================
@app.route('/delete-exam-type', methods=['POST'])
def delete_exam_type():
    if 'user_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401

    data = request.json
    exam_type_id = data.get('id')
    if not exam_type_id:
        return jsonify({'error': 'ID ya mtihani haijabainishwa'}), 400

    try:
        # 1. Pata exam_registration IDs zote za exam_type hii
        exam_regs = supabase.table('exam_registration')\
            .select('id')\
            .eq('exam_type_id', exam_type_id)\
            .execute()
        exam_reg_ids = [reg['id'] for reg in exam_regs.data]

        # 2. Futa marks_table (kwa kutumia exam_registration_ids)
        if exam_reg_ids:
            supabase.table('marks_table')\
                .delete()\
                .in_('exam_registration_id', exam_reg_ids)\
                .execute()

        # 3. Futa exam_registration yenye exam_type_id hii
        supabase.table('exam_registration')\
            .delete()\
            .eq('exam_type_id', exam_type_id)\
            .execute()

        # 4. Futa exam_type yenyewe
        supabase.table('exam_types').delete().eq('id', exam_type_id).execute()

        return jsonify({'message': 'Aina ya mtihani imefutwa pamoja na data zote zinazohusiana.'}), 200
    except Exception as e:
        return jsonify({'error': f'Imeshindwa kufuta: {str(e)}'}), 500

@app.route('/system-usage')
def system_usage():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    try:
        result = supabase.rpc('get_usage_estimates').execute()
        usage = result.data
        return render_template('system_usage.html', usage=usage)
    except Exception as e:
        flash(f'Kosa: {e}', 'danger')
        return redirect(url_for('dashboard'))

@app.route('/export-analysis', methods=['POST'])
def export_analysis():
    try:
        data = request.json
        exam_type_id = data.get('exam_type_id')
        class_name = data.get('class')
        exam_name = data.get('exam_name', 'Mtihani')

        if not exam_type_id or not class_name:
            return jsonify({'error': 'Missing parameters'}), 400

        # Call Supabase RPC
        result = supabase.rpc('get_exam_analysis', {
            'p_exam_type_id': int(exam_type_id),
            'p_class': class_name
        }).execute()

        if not result.data:
            return jsonify({'error': 'Hakuna data ya uchambuzi'}), 404

        analysis = result.data

        # Create workbook
        wb = Workbook()

        # Helper function for styling
        def style_header(cell):
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1e4a3d", end_color="1e4a3d", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Sheet 1: Subject stats
        ws1 = wb.active
        ws1.title = "Subjects performance summary"
        headers1 = ['Subject', 'Registered', 'Abs', 'A', 'B', 'C', 'D', 'F', 'GPA']
        for col, h in enumerate(headers1, 1):
            cell = ws1.cell(row=1, column=col, value=h)
            style_header(cell)

        if analysis.get('subject_stats'):
            for i, subj in enumerate(analysis['subject_stats'], 2):
                ws1.cell(row=i, column=1, value=subj.get('subject', ''))
                ws1.cell(row=i, column=2, value=subj.get('total_enrolled', 0))
                ws1.cell(row=i, column=3, value=subj.get('absent', 0))
                ws1.cell(row=i, column=4, value=subj.get('count_A', 0))
                ws1.cell(row=i, column=5, value=subj.get('count_B', 0))
                ws1.cell(row=i, column=6, value=subj.get('count_C', 0))
                ws1.cell(row=i, column=7, value=subj.get('count_D', 0))
                ws1.cell(row=i, column=8, value=subj.get('count_F', 0))
                ws1.cell(row=i, column=9, value=subj.get('gpa', 0))

        # Adjust column widths
        for col in range(1, len(headers1)+1):
            ws1.column_dimensions[get_column_letter(col)].width = 15
        ws1.freeze_panes = 'A2'

        # Sheet 2: Top best
        ws2 = wb.create_sheet("Wanafunzi Bora")
        headers2 = ['Nafasi', 'Student name', 'Point', 'Marks', 'Division']
        for col, h in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            style_header(cell)
        if analysis.get('top_best'):
            for i, s in enumerate(analysis['top_best'], 2):
                ws2.cell(row=i, column=1, value=s.get('position', ''))
                ws2.cell(row=i, column=2, value=s.get('student_name', ''))
                ws2.cell(row=i, column=3, value=s.get('total_points', 0))
                ws2.cell(row=i, column=4, value=s.get('total_marks', 0))
                ws2.cell(row=i, column=5, value=s.get('division', ''))
        for col in range(1, len(headers2)+1):
            ws2.column_dimensions[get_column_letter(col)].width = 20
        ws2.freeze_panes = 'A2'

        # Sheet 3: Top worst
        ws3 = wb.create_sheet("Waliofeli Zaidi")
        for col, h in enumerate(headers2, 1):
            cell = ws3.cell(row=1, column=col, value=h)
            style_header(cell)
        if analysis.get('top_worst'):
            for i, s in enumerate(analysis['top_worst'], 2):
                ws3.cell(row=i, column=1, value=s.get('position', ''))
                ws3.cell(row=i, column=2, value=s.get('student_name', ''))
                ws3.cell(row=i, column=3, value=s.get('total_points', 0))
                ws3.cell(row=i, column=4, value=s.get('total_marks', 0))
                ws3.cell(row=i, column=5, value=s.get('division', ''))
        for col in range(1, len(headers2)+1):
            ws3.column_dimensions[get_column_letter(col)].width = 20
        ws3.freeze_panes = 'A2'

        # Sheet 4: Divisions
        ws4 = wb.create_sheet("Divisions summary")
        headers4 = ['Sex', 'I', 'II', 'III', 'IV', '0', 'Abs', 'Total']
        for col, h in enumerate(headers4, 1):
            cell = ws4.cell(row=1, column=col, value=h)
            style_header(cell)
        if analysis.get('division_summary'):
            for i, d in enumerate(analysis['division_summary'], 2):
                ws4.cell(row=i, column=1, value=d.get('sex', ''))
                ws4.cell(row=i, column=2, value=d.get('I', 0))
                ws4.cell(row=i, column=3, value=d.get('II', 0))
                ws4.cell(row=i, column=4, value=d.get('III', 0))
                ws4.cell(row=i, column=5, value=d.get('IV', 0))
                ws4.cell(row=i, column=6, value=d.get('0', 0))
                ws4.cell(row=i, column=7, value=d.get('Abs', 0))
                ws4.cell(row=i, column=8, value=d.get('Total', 0))
        for col in range(1, len(headers4)+1):
            ws4.column_dimensions[get_column_letter(col)].width = 12
        ws4.freeze_panes = 'A2'

        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        from datetime import datetime
        now = datetime.now()
        filename = f"Uchambuzi_{class_name}_{exam_name}_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"

        # Kwa Flask 1.x tumia attachment_filename, kwa 2.x tumia download_name
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,   # Flask 2.0+
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        # Tuma kosa kamili kwa mteja (kwa debugging)
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

#===================================================================================================
#                                                         STREMED RESULTS
#=======================================================================================================

@app.route('/stream-results')
def stream_results_page():
    exam_types = supabase.table('exam_types').select('id, exam_name').execute().data
    classes = ['Form1', 'Form2', 'Form3', 'Form4', 'Form5', 'Form6']
    streams = ['A', 'B', 'C', 'D']
    # Orodha ya miaka (miaka 4 iliyopita hadi mwaka ujao)
    current_year = datetime.now().year
    years = list(range(current_year - 3, current_year + 2))
    terms = [1, 2, 3]
    return render_template('stream_results.html',
                           exam_types=exam_types,
                           classes=classes,
                           streams=streams,
                           years=years,
                           terms=terms,
                           current_year=current_year,
                           current_term=1)   # chaguo-msingi muhula 1

@app.route('/api/stream-results', methods=['POST'])
def api_stream_results():
    data = request.json
    exam_type_id = data.get('exam_type_id')
    class_name = data.get('class')
    stream = data.get('stream', 'all')
    year = data.get('year')      # inaweza kuwa None au namba
    term = data.get('term')      # inaweza kuwa None au namba

    if not exam_type_id or not class_name:
        return jsonify({'error': 'Exam type na class vinahitajika'}), 400

    try:
        # Piga Supabase RPC
        result = supabase.rpc('get_stream_results', {
            'p_exam_type_id': int(exam_type_id),
            'p_class': class_name,
            'p_stream': stream,
            'p_year': year,        # Supabase inakubali None kama NULL
            'p_term': term
        }).execute()
        return jsonify(result.data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# -------------------------------------------------------------
def format_index(seq_num, pattern, start):
    """Badilisha muundo wa namba (mfano '{number:04d}') kuwa namba halisi."""
    actual = start + seq_num - 1
    def repl(match):
        spec = match.group(1)
        if spec is None:
            return str(actual)
        if spec == 'd':
            return str(actual)
        if spec.startswith('0'):
            # Ondoa 'd' mwishoni kama ipo
            width_str = spec.rstrip('d')
            width = int(width_str)
            return str(actual).zfill(width)
        # Kama ni namba tu (mfano '{number:4}'), ondoa 'd' kama ipo
        width_str = spec.rstrip('d')
        if width_str.isdigit():
            width = int(width_str)
            return str(actual).zfill(width)
        return str(actual)
    new_pattern = re.sub(r'\{number(?::([^}]+))?\}', repl, pattern)
    return new_pattern

def get_seat_order(rows, columns, door_pos):
    """
    Rudi orodha ya (row, col) kwa mpangilio wa zig-zag kuanzia mlango.
    door_pos: 'top-left', 'top-right', 'bottom-left', 'bottom-right'
    """
    # Amua mwelekeo wa kwanza wa safu mlalo (down = juu→chini, up = chini→juu)
    if door_pos in ('top-left', 'top-right'):
        first_direction = 'down'
    else:
        first_direction = 'up'

    # Amua mwelekeo wa safu wima (kushoto→kulia au kulia→kushoto)
    if door_pos in ('top-left', 'bottom-left'):
        col_range = range(columns)
    else:
        col_range = range(columns-1, -1, -1)

    order = []
    for idx, col in enumerate(col_range):
        # Kila safu wima inabadilisha mwelekeo (zig-zag)
        if idx % 2 == 0:
            direction = first_direction
        else:
            direction = 'up' if first_direction == 'down' else 'down'

        if direction == 'down':
            for row in range(rows):
                order.append((row, col))
        else:  # up
            for row in range(rows-1, -1, -1):
                order.append((row, col))
    return order
@app.route('/download-sitting-plan', methods=['POST'])
def download_sitting_plan():
    data = request.get_json()
    total = int(data.get('total', 40))
    start = int(data.get('start', 1))
    pattern = data.get('pattern', 'S3177/{number:04d}')
    door_pos = data.get('doorPos', 'top-left')
    columns = int(data.get('columns', 5))
    form_text = data.get('formText', 'IV')
    room_text = data.get('roomText', '01')
    heading1 = data.get('heading1', 'KIBINDU SECONDARY SCHOOL')
    heading2 = data.get('heading2', 'MIDTERM EXAMINATION 2026')

    # Tengeneza namba za desk
    desk_numbers = [format_index(i, pattern, start) for i in range(1, total+1)]

    rows = (total + columns - 1) // columns
    seat_order = get_seat_order(rows, columns, door_pos)

    # Unda matrix tupu (None kwa seli tupu)
    matrix = [[None] * columns for _ in range(rows)]
    for idx, (r, c) in enumerate(seat_order):
        if idx < len(desk_numbers):
            matrix[r][c] = desk_numbers[idx]
        # hakuna haja ya kuweka 'else' kwa sababu tayari ni None

    door_class = {
        'top-left': 'door-top-left',
        'top-right': 'door-top-right',
        'bottom-left': 'door-bottom-left',
        'bottom-right': 'door-bottom-right'
    }.get(door_pos, 'door-top-left')

    html_string = render_template('sitting_plan_template.html',
                                  heading1=heading1,
                                  heading2=heading2,
                                  form_text=form_text,
                                  room_text=room_text,
                                  columns=columns,
                                  rows=rows,
                                  matrix=matrix,
                                  door_class=door_class)

    pdf_file = HTML(string=html_string).write_pdf(
        stylesheets=[CSS(string='@page { size: A4; margin: 1.2cm; }')]
    )
    return send_file(BytesIO(pdf_file), as_attachment=True,
                     download_name='sitting_plan.pdf',
                     mimetype='application/pdf')

@app.route('/isal')
def isal():
    classes = ['Form1', 'Form2', 'Form3', 'Form4', 'Form5', 'Form6']
    current_year = datetime.now().year
    years = list(range(current_year - 3, current_year + 1))
    return render_template('isal.html', classes=classes, years=years, current_year=current_year)
@app.route('/api/isal', methods=['POST'])
def api_isal():
    data = request.get_json()
    class_name = data.get('class')
    start_index = data.get('start_index', 1)
    end_index = data.get('end_index')
    year = data.get('year')  # hiari, si lazima

    if not class_name:
        return jsonify({'error': 'Class inahitajika'}), 400
    try:
        start_index = int(start_index)
        if start_index < 1:
            raise ValueError
    except:
        return jsonify({'error': 'Start index lazima iwe namba chanya'}), 400

    try:
        end_index = int(end_index) if end_index else None
        if end_index is not None and end_index < start_index:
            return jsonify({'error': 'End index lazima iwe ≥ start index'}), 400
    except:
        return jsonify({'error': 'End index lazima iwe namba'}), 400

    try:
        result = supabase.rpc('get_index_sheet', {
            'p_class': class_name,
            'p_start_index': start_index,
            'p_end_index': end_index,
            'p_year': year
        }).execute()
        return jsonify(result.data)
    except Exception as e:
        print(f"Supabase RPC error: {e}")
        return jsonify({'error': str(e)}), 500



@app.route('/export-index-sheet', methods=['POST'])
def export_index_sheet():
    try:
        data = request.json
        class_name = data.get('class')
        start_index = data.get('start_index', 1)
        end_index = data.get('end_index')
        year = data.get('year')

        if not class_name:
            return jsonify({'error': 'Class inahitajika'}), 400

        start_index = int(start_index) if start_index else 1
        if start_index < 1:
            return jsonify({'error': 'Start index lazima iwe chanya'}), 400

        if end_index:
            end_index = int(end_index)
            if end_index < start_index:
                return jsonify({'error': 'End index lazima iwe ≥ start index'}), 400
        else:
            end_index = None

        # Pata data kutoka Supabase function
        result = supabase.rpc('get_index_sheet', {
            'p_class': class_name,
            'p_start_index': start_index,
            'p_end_index': end_index,
            'p_year': year
        }).execute()
        students = result.data

        if not students:
            return jsonify({'error': 'Hakuna data ya kuexport'}), 404

        # Unda Excel
        wb = Workbook()
        ws = wb.active
        ws.title = f"{class_name}-{start_index}"

        # Header (kwa mitindo)
        from openpyxl.styles import Font, PatternFill, Alignment
        headers = ['Index', 'Jina la Mwanafunzi', 'Jinsia']
        ws.append(headers)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1e4a3d", end_color="1e4a3d", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        for student in students:
            ws.append([
                student.get('index_number', ''),
                student.get('student_name', ''),
                student.get('sex', '')
            ])

        # Adjust column widths
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 30)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"IndexSheet_{class_name}_{start_index}_{end_index or 'all'}.xlsx"
        return send_file(output, as_attachment=True, download_name=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        # Tafadhali andika kwenye logi
        print(f"Export error: {e}")
        return jsonify({'error': str(e)}), 500


       #============================ DOWNLOAD FULL CLASS RESULT ============================
@app.route('/download-results')
def download_results_page():
    classes = ['Form1', 'Form2', 'Form3', 'Form4', 'Form5', 'Form6']
    return render_template('download_results.html', classes=classes)


@app.route('/api/exam-types')
def get_exam_types():
    try:
        result = supabase.table('exam_types').select('id, exam_name').order('exam_name').execute()
        return jsonify(result.data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/export-class-full-results', methods=['POST'])
def export_class_full_results():
    data = request.json
    exam_type_id = data.get('exam_type_id')
    class_name = data.get('class')
    if not exam_type_id or not class_name:
        return jsonify({'error': 'Exam type and class required'}), 400

    try:
        # Pata data kutoka kwa Supabase function
        result = supabase.rpc('get_class_full_results', {
            'p_exam_type_id': int(exam_type_id),
            'p_class': class_name
        }).execute()
        payload = result.data  # { subjects_list: [...], students: [...] }

        subjects_list = payload.get('subjects_list', [])
        students = payload.get('students', [])

        if not students:
            return jsonify({'error': 'Hakuna matokeo ya kuexport'}), 404

        # Unda workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Matokeo_{class_name}"

        # Header (safu wima)
        headers = ['Jina la Mwanafunzi', 'Jinsia'] + subjects_list + ['Wastani', 'Pointi', 'Division']
        ws.append(headers)

        # Styling ya header
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1e4a3d", end_color="1e4a3d", fill_type="solid")
        header_align = Alignment(horizontal="center")
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # Data rows
        for student in students:
            row_data = [
                student.get('student_name', ''),
                student.get('sex', '')
            ]
            # Kwa kila somo kwenye subjects_list, toa marks+grade
            for sub in subjects_list:
                subj_info = student.get('subjects', {}).get(sub, {})
                marks = subj_info.get('marks')
                grade = subj_info.get('grade', '')
                if marks is not None:
                    row_data.append(f"{marks}")
                else:
                    row_data.append('-')
            # Wastani, Pointi, Division
            avg = student.get('average')
            row_data.append(round(avg, 2) if avg is not None else '-')
            points = student.get('total_points')
            if points == 'inc':
                row_data.append('inc')
            else:
                row_data.append(points if points is not None else '-')
            row_data.append(student.get('division') or '-')
            ws.append(row_data)

        # Rekebisha upana wa safu wima
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 35)

        # Hifadhi kwenye BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Tuma faili kwa mteja
        filename = f"Matokeo_{class_name}_exam_{exam_type_id}.xlsx"
        return send_file(output, as_attachment=True, download_name=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        print(f"Export error: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/class-full-results', methods=['POST'])
def class_full_results():
    data = request.json
    exam_type_id = data.get('exam_type_id')
    class_name = data.get('class')
    if not exam_type_id or not class_name:
        return jsonify({'error': 'Exam type and class required'}), 400
    try:
        result = supabase.rpc('get_class_full_results', {
            'p_exam_type_id': int(exam_type_id),
            'p_class': class_name
        }).execute()
        return jsonify(result.data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


import zipfile


@app.route('/download-all-pdfs', methods=['POST'])
def download_all_pdfs():
    try:
        data = request.get_json()
        exam_type_id = data.get('exam_type_id')
        class_name = data.get('class')
        if not exam_type_id or not class_name:
            return jsonify({'error': 'Exam type and class required'}), 400

        # Pata data ya darasa zima (kutumia Supabase function)
        result = supabase.rpc('get_class_full_results', {
            'p_exam_type_id': int(exam_type_id),
            'p_class': class_name
        }).execute()
        payload = result.data
        students = payload.get('students', [])
        subjects_list = payload.get('subjects_list', [])

        if not students:
            return jsonify({'error': 'Hakuna wanafunzi'}), 404

        # Unda ZIP kwenye memory
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for student in students:
                # Tengeneza HTML kwa kutumia template ya PDF
                html_string = render_template('student_report_pdf.html',
                                              student=student,
                                              subjects_list=subjects_list,
                                              exam_type_id=exam_type_id,
                                              class_name=class_name)
                pdf_data = HTML(string=html_string).write_pdf(
                    stylesheets=[CSS(string='@page { size: A4; margin: 1.5cm; }')]
                )
                # Jina la faili (ondoa nafasi na herufi zisizo salama)
                safe_name = student['student_name'].replace(' ', '_').replace('/', '_')
                filename = f"{safe_name}.pdf"
                zip_file.writestr(filename, pdf_data)

        zip_buffer.seek(0)
        return send_file(zip_buffer, as_attachment=True,
                         download_name=f"Matokeo_{class_name}_exam_{exam_type_id}.zip",
                         mimetype='application/zip')
    except Exception as e:
        print(f"ZIP creation error: {e}")
        return jsonify({'error': str(e)}), 500

    #========================================================================
@app.route('/student-pdf', methods=['POST'])
def student_pdf():
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data received'}), 400

        student_id = data.get('student_id')
        exam_type_id = data.get('exam_type_id')
        class_name = data.get('class')

        if not student_id or not exam_type_id or not class_name:
            return jsonify({'error': 'Missing parameters: student_id, exam_type_id, class'}), 400

        # Pata matokeo ya darasa zima
        result = supabase.rpc('get_class_full_results', {
            'p_exam_type_id': int(exam_type_id),
            'p_class': class_name
        }).execute()
        payload = result.data
        students = payload.get('students', [])
        student_data = next((s for s in students if s['student_id'] == student_id), None)
        if not student_data:
            return jsonify({'error': f'Student with id {student_id} not found'}), 404

        subjects_list = payload.get('subjects_list', [])

        # Angalia ikiwa template ipo
        try:
            html_string = render_template('student_report_pdf.html',
                                          student=student_data,
                                          subjects_list=subjects_list,
                                          exam_type_id=exam_type_id,
                                          class_name=class_name)
        except Exception as e:
            return jsonify({'error': f'Template error: {str(e)}'}), 500

        # Geuza kuwa PDF
        try:
            pdf_file = HTML(string=html_string).write_pdf(
                stylesheets=[CSS(string='@page { size: A4; margin: 1.5cm; }')]
            )
        except Exception as e:
            return jsonify({'error': f'PDF generation error: {str(e)}'}), 500

        filename = f"Report_{student_data['student_name'].replace(' ', '_')}.pdf"
        return send_file(BytesIO(pdf_file), as_attachment=True,
                         download_name=filename,
                         mimetype='application/pdf')

    except Exception as e:
        # Hapa ndipo tunahakikisha kosa lolote linarejeshwa kama JSON
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/public-download-joining', methods=['GET'])
def public_download_joining():
    form = request.args.get('form')  # kwa mfano 'Form1', hiari
    try:
        # Tafuta faili la hivi karibuni kwa kidato (au zote)
        query = supabase.table('join_files').select('id, file_path, file_name')
        if form:
            query = query.eq('form', form)
        result = query.order('created_at', desc=True).limit(1).execute()
        if not result.data:
            return jsonify({'error': 'Hakuna joining instruction iliyopatikana'}), 404

        file_path = result.data[0]['file_path']
        # Unda signed URL yenye muda wa sekunde 300 (dakika 5)
        signed_url = supabase.storage.from_('join_files').create_signed_url(file_path, 300)
        if not signed_url:
            return jsonify({'error': 'Imeshindwa kuunda signed URL'}), 500

        # Rudisha redirection kwenye signed URL (mzazi atapakua moja kwa moja)
        return redirect(signed_url['signedURL'])
    except Exception as e:
        return jsonify({'error': str(e)}), 500



# Hii ndio njia ya kutoa sitemap.xml
@app.route("/sitemap.xml")
def sitemap():
    return sitemapper.generate()

@app.route('/robots.txt')
def robots_txt():
    return send_from_directory(app.static_folder, 'robots.txt', mimetype='text/plain')

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)